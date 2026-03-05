"""
pdf_a_excel_EOI.py
Procesa formularios PDF de los programas EOI:
  - "Generación Digital PYMES: Personas de Equipos Directivos"  → formulario.xlsm
  - "Agentes del Cambio" (modelo _1_, pyme 10-249)              → _1_FORMULARIO_AGENTES...xlsm
  - "Agentes del Cambio" (modelo _2_, pyme 1-249)               → _2_FORMULARIO_AGENTES...xlsm

Las plantillas Excel están embebidas en el propio script (base64).
El script detecta automáticamente el tipo de PDF y no pide plantillas al usuario.

Dependencias: pip install pdfplumber pypdf openpyxl
"""

import os
import re
import base64
import tempfile
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime

import pdfplumber
import openpyxl
from pypdf import PdfReader

# Importar las plantillas en base64 (fichero generado aparte)
from plantillas_b64 import PLANTILLA_DIRECTIVOS_B64, PLANTILLA_AGENTES1_B64, PLANTILLA_AGENTES2_B64

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ──────────────────────────────────────────────────────────────────────────────

SEP_X = 350  # Separador X entre etiquetas y valores en el PDF

TIPO_DIRECTIVOS = "directivos"
TIPO_AGENTES_1  = "agentes_1"   # pyme 10-249
TIPO_AGENTES_2  = "agentes_2"   # pyme 1-249

HOJA_POR_TIPO = {
    TIPO_DIRECTIVOS: "Formulario Directivos",
    TIPO_AGENTES_1:  "formulario agentes",
    TIPO_AGENTES_2:  "formulario agentes",
}

# Filas de declaraciones (se rellenan con "Acepto" automáticamente)
DECLARACIONES_POR_TIPO = {
    TIPO_DIRECTIVOS: [85, 86, 87, 88, 89, 90, 91],
    TIPO_AGENTES_1:  [80, 81, 82, 83, 84, 85],
    TIPO_AGENTES_2:  [79, 80, 81, 82, 83, 84, 85],
}

# Plantillas embebidas (base64 → bytes en tiempo de ejecución)
PLANTILLA_B64_POR_TIPO = {
    TIPO_DIRECTIVOS: PLANTILLA_DIRECTIVOS_B64,
    TIPO_AGENTES_1:  PLANTILLA_AGENTES1_B64,
    TIPO_AGENTES_2:  PLANTILLA_AGENTES2_B64,
}

# ──────────────────────────────────────────────────────────────────────────────
# MAPEOS PDF → FILA EXCEL (col B)
# ──────────────────────────────────────────────────────────────────────────────

MAPEO_DIRECTIVOS = {
    "PRIMER APELLIDO":          6,
    "SEGUNDO APELLIDO":         7,
    "NOMBRE":                   8,
    "TIPO DE DOCUMENTO (NIF, NIE, DOCUMENTO ID, PASAPORTE)": 9,
    "No DE DOCUMENTO":          10,
    "SEXO (M-F-NB)":            11,
    "FECHA DE NACIMIENTO":      12,
    "DIRECCION":                13,
    "CIUDAD":                   14,
    "CODIGO POSTAL":            15,
    "CCAA":                     16,
    "PROVINCIA":                17,
    "TELEFONO":                 18,
    "EMAIL":                    19,
    "RESIDE EN UNA LOCALIDAD CON UN NUMERO DE HABITANTES INFERIOR A 5.000": 20,
    "PERSONA CON DISCAPACIDAD": 21,
    "NIVEL DE ESTUDIOS":        22,
    "TITULACION":               23,
    "NOMBRE EMPRESA (RAZON SOCIAL)": 26,
    "RELACION CON LA EMPRESA":  27,
    "DEPARTAMENTO":             28,
    "PUESTO/CARGO":             29,
    "NIF EMPRESA":              32,
    "ACTIVIDAD DE LA EMPRESA (CODIGOS CNAE NIVEL LETRA)": 33,
    "TAMANO EMPRESA":           34,
    "DIRECCION (EMPRESA)":      35,
    "CIUDAD (EMPRESA)":         36,
    "CODIGO POSTAL (EMPRESA)":  37,
    "CCAA (EMPRESA)":           38,
    "PROVINCIA (EMPRESA)":      39,
    "TELEFONO (EMPRESA)":       40,
    "PAGINA WEB (EMPRESA)":     41,
    "ANTIGUEDAD DE LA EMPRESA": 42,
    "FACTURACION ULTIMO ANO":   43,
    "AMBITO RURAL (SI EL CENTRO DE TRABAJO SE SITUA EN UN MUNICIPIO": 44,
    "NIVEL DE MADUREZ DIGITAL DE LA EMPRESA EN EL": 45,
    "CANALES DE RELACION DE LA EMPRESA CON CLIENTES Y PROVEEDORES": 46,
    "PROFESIONALES CON PERFIL TIC EN LA EMPRESA": 47,
    "EMPRESA CON POLITICAS DE SOSTENIBILIDAD": 48,
    "EMPRESA CON POLITICAS O PLANES DE TRANSFORMACION DIGITAL": 49,
    "LA MAXIMA RESPONSABLE DE LA EMPRESA O MAS DEL 50% DEL EQUIPO DIRECTIVO": 50,
    "PORCENTAJE DE MUJERES CON RELACION LABORAL CON LA EMPRESA": 51,
    "DESCRIBIR MOTIVACION PARA CURSAR EL PROGRAMA": 54,
}

MAPEO_AGENTES = {
    "PRIMER APELLIDO":          6,
    "SEGUNDO APELLIDO":         7,
    "NOMBRE":                   8,
    "TIPO DE DOCUMENTO (NIF, NIE, DOCUMENTO ID, PASAPORTE)": 9,
    "No DE DOCUMENTO":          10,
    "SEXO (M-F-NB)":            11,
    "FECHA DE NACIMIENTO":      12,
    "DIRECCION":                13,
    "CIUDAD":                   14,
    "CODIGO POSTAL":            15,
    "CCAA":                     16,
    "PROVINCIA":                17,
    "TELEFONO":                 18,
    "EMAIL":                    19,
    "RESIDE EN UNA LOCALIDAD CON UN NUMERO DE HABITANTES INFERIOR A 5": 20,
    "PERSONA CON DISCAPACIDAD": 21,
    "PERFIL DE LINKEDIN":       22,
    "NIVEL DE ESTUDIOS":        25,
    "FORMACION COMPLEMENTARIA EN DIGITALIZACION": 26,
    "FORMACION COMPLEMENTARIA EN GESTION DE PROYECTOS": 27,
    "ANOS DE EXPERIENCIA LABORAL": 30,
    "EXPERIENCIA LABORAL EN PUESTOS DE DIGITALIZACION": 31,
    "SITUACION LABORAL ACTUAL": 32,
    "NOMBRE EMPRESA (RAZON SOCIAL)": 34,
    "NIF":                      35,
    "DEPARTAMENTO (EMPRESA)":   36,
    "PUESTO/CARGO (EMPRESA)":   37,
    "ACTIVIDAD DE LA EMPRESA (CODIGOS CNAE NIVEL LETRA)": 38,
    "TAMANO EMPRESA":           39,
    "DIRECCION (EMPRESA)":      40,
    "CIUDAD (EMPRESA)":         41,
    "CODIGO POSTAL (EMPRESA)":  42,
    "CCAA (EMPRESA)":           43,
    "PROVINCIA (EMPRESA)":      44,
    "TELEFONO (EMPRESA)":       45,
    "PAGINA WEB (EMPRESA)":     46,
    "ANTIGUEDAD DE LA EMPRESA": 47,
    "FACTURACION ULTIMO ANO":   48,
    "AMBITO RURAL":             49,
    "NIVEL DE MADUREZ DIGITAL DE LA EMPRESA EN EL": 50,
    "EMPRESA CON POLITICAS DE SOSTENIBILIDAD": 51,
    "EMPRESA CON POLITICAS O PLANES DE TRANSFORMACION DIGITAL": 52,
    "LA MAXIMA RESPONSABLE DE LA EMPRESA O MAS DEL 50% DEL EQUIPO DIRECTIVO": 53,
    "PORCENTAJE DE MUJERES CON RELACION LABORAL CON LA EMPRESA": 54,
    "MOTIVACION PARA CURSAR EL PROGRAMA": 57,
}

MAPEO_POR_TIPO = {
    TIPO_DIRECTIVOS: MAPEO_DIRECTIVOS,
    TIPO_AGENTES_1:  MAPEO_AGENTES,
    TIPO_AGENTES_2:  MAPEO_AGENTES,
}

# Filas donde el valor debe ser SI/NO
FILAS_SI_NO_POR_TIPO = {
    TIPO_DIRECTIVOS: {20, 21, 44, 48, 49, 50},
    TIPO_AGENTES_1:  {20, 21, 49, 51, 52, 53},
    TIPO_AGENTES_2:  {20, 21, 49, 51, 52, 53},
}

# ──────────────────────────────────────────────────────────────────────────────
# UTILIDADES
# ──────────────────────────────────────────────────────────────────────────────

def normalizar(texto: str) -> str:
    """Convierte a mayúsculas, elimina tildes y colapsa espacios."""
    if not texto:
        return ""
    # Ambas cadenas de exactamente la misma longitud
    origen  = "áéíóúàèìòùâêîôûäëïöüÁÉÍÓÚÀÈÌÒÙÂÊÎÔÛÄËÏÖÜñÑ"
    destino = "aeiouaeiouaeiouaeiouAEIOUAEIOUAEIOUAEIOUnN"
    assert len(origen) == len(destino), f"tabla rota: {len(origen)} vs {len(destino)}"
    tabla = str.maketrans(origen, destino)
    # Eliminar también el carácter 'º' que no tiene equivalente directo
    resultado = texto.upper().translate(tabla).replace("º", "")
    return " ".join(resultado.split())


def _extraer_si_no(valor: str) -> str:
    """
    Extrae el SI o NO que representa la respuesta real del alumno.
    Los campos del PDF tienen formato "...opciones... SI / NO RESPUESTA",
    por eso buscamos el último SI o NO de la cadena.
    """
    matches = list(re.finditer(r'\b(SI|NO)\b', str(valor).strip(), re.IGNORECASE))
    return matches[-1].group(1).upper() if matches else valor


def _limpiar_facturacion(valor: str) -> str:
    """
    Cuando pdfplumber concatena etiqueta+valor, el campo tiene formato:
    "0 - 500.000€ / 500.000 - 1M€ / ... / + 4 M€ 0 - 500.000€"
    Extraemos la última opción (la que ha marcado el alumno).
    """
    valor = str(valor).strip()
    # Buscar la última ocurrencia de un rango de facturación
    m = re.search(r'([+\d][^/€]*€)\s*$', valor)
    if m:
        return m.group(1).strip()
    # Fallback: último token que empiece por dígito o '+'
    partes = valor.split()
    for i in range(len(partes) - 1, -1, -1):
        if re.match(r'^[\d+]', partes[i]):
            return " ".join(partes[i:])
    return valor


def _limpiar_opcion_lista(valor: str) -> str:
    """
    Limpia campos tipo lista de opciones donde pdfplumber concatena la lista + respuesta.
    Ejemplos:
      "últimos 5 años / 5 - 10 años / + 10 años 5 - 10 años"
        → "5 - 10 años"
      "1 – 9 trabajadores / 10 - 49 trabajadores / 50 - 249 trabajadores 1 - 9 trabajadores"
        → "1 - 9 trabajadores"
      "0 - 500.000€ / 500.000 - 1M€ / ... / + 4 M€ 0 - 500.000€"
        → "0 - 500.000€"
    Estrategia: la respuesta aparece al final, tras la última opción.
    Se extrae buscando la opción más corta que termine el string.
    """
    valor = str(valor).strip()
    if '/' not in valor:
        return valor

    # Normalizar guiones (em-dash → hyphen) para comparación uniforme
    valor_norm = valor.replace('–', '-').replace('—', '-')

    # Obtener todas las opciones separadas por '/'
    opciones = [p.strip() for p in valor_norm.split('/')]
    # La respuesta está al final de la última "opción" (que en realidad es "ultima_opcion RESPUESTA")
    ultima = opciones[-1]

    # Buscar si alguna opción anterior aparece al inicio de `ultima`
    # Si es así, lo que queda tras esa opción es la respuesta
    for opcion in opciones[:-1]:
        if ultima.startswith(opcion):
            respuesta = ultima[len(opcion):].strip()
            if respuesta:
                # Restaurar el valor original con la respuesta encontrada
                # Buscar en el valor original para devolver con formato original
                return respuesta.replace('-', valor[valor.find(respuesta[0]):valor.find(respuesta[0])+2] if valor.find(respuesta[0]) >= 0 else '-')

    # Buscar si alguna opción está contenida completa al FINAL de ultima
    for opcion in opciones:
        opcion_s = opcion.strip()
        if opcion_s and ultima.endswith(opcion_s):
            return opcion_s

    # Fallback: devolver la última parte tal cual
    return ultima


def _limpiar_madurez(valor: str) -> str:
    niveles = ("bajo", "medio", "alto", "muy", "basico", "básico")
    partes = str(valor).split()
    for i, p in enumerate(partes):
        if p.lower() in niveles:
            return " ".join(partes[i:])
    return valor


def parsear_fecha(texto: str):
    """Parsea fechas en formatos: DD/MM/YYYY, DD-MM-YYYY, YYYY-MM-DD y texto con mes en español."""
    meses = {
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
        "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
        "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    }
    texto_str = str(texto).strip()

    # DD/MM/YYYY o DD-MM-YYYY
    m = re.match(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', texto_str)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    # YYYY-MM-DD
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', texto_str)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    # "12 de enero de 1990" o "12 enero 1990"
    texto_lower = texto_str.lower()
    for nombre, num in meses.items():
        if nombre in texto_lower:
            partes = re.findall(r'\d+', texto_lower)
            if len(partes) >= 2:
                try:
                    return datetime(int(partes[-1]), num, int(partes[0]))
                except ValueError:
                    pass

    return texto  # Devolver original si no se puede parsear


# ──────────────────────────────────────────────────────────────────────────────
# GESTIÓN DE PLANTILLAS EMBEBIDAS
# ──────────────────────────────────────────────────────────────────────────────

def obtener_ruta_plantilla(tipo: str) -> str:
    """
    Decodifica la plantilla embebida en base64 y la escribe en un fichero
    temporal. Devuelve la ruta al fichero temporal.
    El llamador es responsable de borrar el fichero cuando termine.
    """
    b64_data = PLANTILLA_B64_POR_TIPO[tipo]
    # b64_data puede ser una tupla de strings (concatenación) o un string
    if isinstance(b64_data, tuple):
        b64_data = "".join(b64_data)
    raw = base64.b64decode(b64_data)
    fd, ruta = tempfile.mkstemp(suffix=".xlsm")
    with os.fdopen(fd, "wb") as f:
        f.write(raw)
    return ruta


# ──────────────────────────────────────────────────────────────────────────────
# DETECCIÓN DEL TIPO DE FORMULARIO
# ──────────────────────────────────────────────────────────────────────────────

def detectar_tipo_pdf(lineas: list) -> str:
    """
    Recibe las líneas ya extraídas del PDF para no releer el fichero.
    Devuelve el tipo de formulario detectado.
    """
    cabecera = " ".join(lineas[:5]).upper()

    if "AGENTES DEL CAMBIO" in cabecera:
        texto_completo = " ".join(lineas).upper()
        if "1-249" in texto_completo or "1 - 249" in texto_completo:
            return TIPO_AGENTES_2
        return TIPO_AGENTES_1

    return TIPO_DIRECTIVOS


# ──────────────────────────────────────────────────────────────────────────────
# EXTRACCIÓN DE CAMPOS
# ──────────────────────────────────────────────────────────────────────────────

def leer_lineas_pdf(ruta_pdf: str) -> list:
    """
    Extrae líneas de TODAS las páginas usando pypdf.
    Fallback a pdfplumber si pypdf falla o devuelve vacío.
    """
    errores = []

    # 1) Intento con pypdf (más tolerante en muchos casos)
    try:
        with open(ruta_pdf, "rb") as fh:
            reader = PdfReader(fh)
            trozos = []
            for page in reader.pages:
                try:
                    trozos.append(page.extract_text() or "")
                except Exception as e:
                    errores.append(f"pypdf(page): {type(e).__name__}: {e}")
            texto = "\n".join(trozos)
        lineas = [l.strip() for l in texto.splitlines() if l.strip()]
        if lineas:
            return lineas
        errores.append("pypdf: sin texto extraible")
    except Exception as e:
        errores.append(f"pypdf: {type(e).__name__}: {e}")

    # 2) Fallback con pdfplumber
    try:
        with pdfplumber.open(ruta_pdf) as pdf:
            texto = "\n".join((p.extract_text() or "") for p in pdf.pages)
        lineas = [l.strip() for l in texto.splitlines() if l.strip()]
        if lineas:
            return lineas
        errores.append("pdfplumber: sin texto extraible")
    except Exception as e:
        errores.append(f"pdfplumber: {type(e).__name__}: {e}")

    raise ValueError("No se pudo extraer texto del PDF. " + " | ".join(errores))


ENCABEZADOS_STOP = {
    "DATOS", "NOMBRE", "NIF", "RELACIÓN", "DEPARTAMENTO", "PUESTO",
    "ACTIVIDAD", "TAMAÑO", "DIRECCIÓN", "CIUDAD", "CODIGO", "CCAA",
    "PROVINCIA", "TELÉFONO", "PAGINA", "ANTIGÜEDAD", "FACTURACIÓN",
    "AMBITO", "NIVEL", "CANALES", "PROFESIONALES", "EMPRESA",
    "LA MÁXIMA", "PORCENTAJE", "SEGUNDO", "TIPO", "Nº", "SEXO",
    "FECHA", "EMAIL", "RESIDE", "PERSONA", "TITULACION", "Describir",
    "PRIMER", "DOCUMENTACIÓN", "ACEPTO", "Declaro", "AUTORIZO",
    "Acepto", "Firma", "AÑOS", "EXPERIENCIA", "SITUACIÓN", "MOTIVACIÓN",
    "Formación", "Perfil",
}


def extraer_linea(etiqueta: str, lineas: list, multilinea: bool = False) -> str:
    patron = re.compile(r'^' + re.escape(etiqueta) + r'\s*(.*)', re.IGNORECASE)
    for i, linea in enumerate(lineas):
        m = patron.match(linea)
        if m:
            valor = m.group(1).strip()
            if multilinea and i + 1 < len(lineas):
                sig = lineas[i + 1]
                if not any(sig.startswith(k) for k in ENCABEZADOS_STOP):
                    valor = (valor + " " + sig).strip()
            return valor
    return ""


def extraer_relacion_empresa_lineas(lineas: list) -> str:
    STOP = {
        "DATOS", "NOMBRE", "NIF", "DEPARTAMENTO", "PUESTO", "ACTIVIDAD",
        "TAMAÑO", "DIRECCIÓN", "CIUDAD", "CODIGO", "CCAA", "PROVINCIA",
        "TELÉFONO", "PAGINA", "ANTIGÜEDAD", "FACTURACIÓN", "AMBITO",
        "NIVEL", "CANALES", "PROFESIONALES", "EMPRESA", "LA MÁXIMA",
        "PORCENTAJE",
    }
    for i, l in enumerate(lineas):
        if "RELACIÓN CON LA EMPRESA" in l:
            valor_en_linea = re.sub(r'RELACI[OÓ]N CON LA EMPRESA\s*', '', l, flags=re.IGNORECASE).strip()
            parte_anterior = ""
            if not valor_en_linea and i > 0 and not any(lineas[i-1].startswith(k) for k in STOP):
                parte_anterior = lineas[i-1].strip()
            partes_post = []
            j = i + 1
            while j < len(lineas):
                if any(lineas[j].startswith(k) for k in STOP):
                    break
                partes_post.append(lineas[j])
                j += 1
            partes = list(filter(None, [parte_anterior, valor_en_linea] + partes_post))
            return " ".join(partes).strip()
    return ""


def extraer_porcentaje_mujeres_lineas(lineas: list) -> str:
    STOP_PORC = {"DATOS", "MOTIVACIÓN", "DOCUMENTACIÓN", "ACEPTO", "Declaro", "Firma", "CONDICIONADO"}
    OPCIONES   = re.compile(r'inferior a 30%|entre 30%\s*y\s*50%|superior a 50%', re.IGNORECASE)

    def _normalizar_opcion(texto: str) -> str:
        """Intenta completar fragmentos como 'a 50%' → 'superior a 50%'."""
        t = texto.strip()
        m = OPCIONES.search(t)
        if m:
            return m.group(0)
        tl = t.lower()
        if "superior" in tl or tl.startswith("a 50"):
            return "superior a 50%"
        if "inferior" in tl or tl.startswith("a 30"):
            return "inferior a 30%"
        if "entre" in tl or ("30%" in tl and "50%" in tl):
            return "entre 30% y 50%"
        return t

    for i, l in enumerate(lineas):
        if "PORCENTAJE DE MUJERES" in l.upper():
            resto = re.sub(
                r'^PORCENTAJE DE MUJERES CON RELACI[OÓ]N LABORAL CON LA EMPRESA[:\s]*',
                '', l, flags=re.IGNORECASE
            ).strip()
            # Quitar lista de opciones si la incluye
            resto = re.sub(
                r'(inferior a 30%\s*/\s*entre 30%\s*y\s*50%\s*/\s*superior a 50%)\s*',
                '', resto, flags=re.IGNORECASE
            ).strip()
            if resto:
                return _normalizar_opcion(resto)
            # Buscar en las líneas siguientes
            for j in range(i + 1, min(i + 4, len(lineas))):
                sig = lineas[j].strip()
                if any(sig.upper().startswith(k.upper()) for k in STOP_PORC):
                    break
                sig_clean = re.sub(
                    r'(inferior a 30%\s*/\s*entre 30%\s*y\s*50%\s*/\s*superior a 50%)\s*',
                    '', sig, flags=re.IGNORECASE
                ).strip()
                if sig_clean:
                    return _normalizar_opcion(sig_clean)
    return ""


def leer_campos_pdf(ruta_pdf: str, lineas: list) -> dict:
    """
    Extrae campos del PDF combinando dos estrategias.
    Recibe las líneas ya leídas para evitar releer el fichero.

    Estrategia 1: coordenadas X (pdfplumber) — todas las páginas.
    Estrategia 2: líneas de texto (ya proporcionadas).
    La estrategia 2 solo sobreescribe campos ausentes en la 1,
    salvo que la 1 haya fallado (pdf_corrompido).
    """
    campos = {}

    # ── Estrategia 1: coordenadas X (todas las páginas) ───────────────────────
    campos_coord = {}
    pdf_corrompido = False
    try:
        with pdfplumber.open(ruta_pdf) as pdf:
            for page in pdf.pages:          # ← todas las páginas
                palabras = page.extract_words()
                filas: dict[int, list] = {}
                for w in palabras:
                    clave_fila = round(w["top"] / 4) * 4
                    filas.setdefault(clave_fila, []).append(w)

                etiqueta_pendiente = ""
                for top, palabras_fila in sorted(filas.items()):
                    etiqueta_partes = [w["text"] for w in palabras_fila if w["x0"] < SEP_X]
                    valor_partes    = [w["text"] for w in palabras_fila if w["x0"] >= SEP_X]
                    etiqueta_raw = " ".join(etiqueta_partes).strip()
                    valor_raw    = " ".join(valor_partes).strip()

                    if etiqueta_raw and valor_raw:
                        campos_coord[normalizar(etiqueta_raw)] = valor_raw
                        etiqueta_pendiente = ""
                    elif etiqueta_raw and not valor_raw:
                        etiqueta_pendiente = etiqueta_raw
                    elif not etiqueta_raw and valor_raw and etiqueta_pendiente:
                        campos_coord[normalizar(etiqueta_pendiente)] = valor_raw
                        etiqueta_pendiente = ""

        # Normalizar espacios en valores de pdfplumber
        for k in list(campos_coord.keys()):
            if isinstance(campos_coord[k], str):
                campos_coord[k] = re.sub(r'\s+', ' ', campos_coord[k]).strip()

        # Verificar calidad: ¿extrajo campos útiles?
        campos_con_valor = sum(1 for v in campos_coord.values() if v and len(str(v)) > 1)
        apellido_coord   = campos_coord.get(normalizar("PRIMER APELLIDO"), "")
        texto_corrupto   = apellido_coord and (
            len(apellido_coord) > 25 or
            sum(1 for c in apellido_coord if not c.isalpha() and c != ' ') > 3
        )
        if campos_con_valor < 5 or texto_corrupto:
            pdf_corrompido = True
            print(f"  ⚠️  pdfplumber: solo {campos_con_valor} campos útiles → usando solo pypdf.")
        else:
            campos.update(campos_coord)

    except Exception as e:
        pdf_corrompido = True
        print(f"  ⚠️  Estrategia coordenadas falló: {e}")

    # ── Estrategia 2: líneas de texto ─────────────────────────────────────────
    # Lista de (etiqueta_en_pdf, clave_normalizada_destino)
    campos_linea = [
        ("PRIMER APELLIDO",           "PRIMER APELLIDO"),
        ("SEGUNDO APELLIDO",          "SEGUNDO APELLIDO"),
        ("NOMBRE",                    "NOMBRE"),
        ("TIPO DE DOCUMENTO (NIF, NIE, Documento ID, Pasaporte)", "TIPO DE DOCUMENTO (NIF, NIE, DOCUMENTO ID, PASAPORTE)"),
        ("Nº de Documento",           "No DE DOCUMENTO"),
        ("SEXO (M-F-NB)",             "SEXO (M-F-NB)"),
        ("FECHA DE NACIMIENTO",       "FECHA DE NACIMIENTO"),
        ("DIRECCION",                 "DIRECCION"),
        ("CIUDAD",                    "CIUDAD"),
        ("CODIGO POSTAL",             "CODIGO POSTAL"),
        ("CCAA",                      "CCAA"),
        ("PROVINCIA",                 "PROVINCIA"),
        ("TELÉFONO",                  "TELEFONO"),
        ("EMAIL",                     "EMAIL"),
        ("Reside en una localidad con un número de habitantes inferior a 5", "RESIDE EN UNA LOCALIDAD CON UN NUMERO DE HABITANTES INFERIOR A 5"),
        ("Persona con discapacidad",  "PERSONA CON DISCAPACIDAD"),
        ("Perfil de LinkedIn",        "PERFIL DE LINKEDIN"),
        ("NIVEL DE ESTUDIOS:",        "NIVEL DE ESTUDIOS"),
        ("Formación complementaria en digitalización (de más de 100 horas)", "FORMACION COMPLEMENTARIA EN DIGITALIZACION"),
        ("Formación complementaria en gestión de proyectos, innovación (de más de 100 horas)", "FORMACION COMPLEMENTARIA EN GESTION DE PROYECTOS"),
        ("AÑOS DE EXPERIENCIA LABORAL:", "ANOS DE EXPERIENCIA LABORAL"),
        ("EXPERIENCIA LABORAL EN PUESTOS DE DIGITALIZACION", "EXPERIENCIA LABORAL EN PUESTOS DE DIGITALIZACION"),
        ("SITUACIÓN LABORAL ACTUAL:", "SITUACION LABORAL ACTUAL"),
        ("NOMBRE EMPRESA (razón social)", "NOMBRE EMPRESA (RAZON SOCIAL)"),
        ("NIF",                       "NIF"),
        ("DEPARTAMENTO (empresa)",    "DEPARTAMENTO (EMPRESA)"),
        ("PUESTO/CARGO (empresa)",    "PUESTO/CARGO (EMPRESA)"),
        ("TAMAÑO EMPRESA:",           "TAMANO EMPRESA"),
        ("DIRECCIÓN (empresa)",       "DIRECCION (EMPRESA)"),
        ("CIUDAD (empresa)",          "CIUDAD (EMPRESA)"),
        ("CODIGO POSTAL (empresa)",   "CODIGO POSTAL (EMPRESA)"),
        ("CCAA (empresa)",            "CCAA (EMPRESA)"),
        ("PROVINCIA (empresa)",       "PROVINCIA (EMPRESA)"),
        ("TELÉFONO (empresa)",        "TELEFONO (EMPRESA)"),
        ("PAGINA WEB (empresa)",      "PAGINA WEB (EMPRESA)"),
        ("ANTIGÜEDAD DE LA EMPRESA:", "ANTIGUEDAD DE LA EMPRESA"),
        ("FACTURACIÓN ÚLTIMO AÑO:",   "FACTURACION ULTIMO ANO"),
        ("AMBITO RURAL",              "AMBITO RURAL"),
        ('NIVEL DE MADUREZ DIGITAL DE LA EMPRESA en el "Test de Diagnóstico Digital" de acelerapyme', "NIVEL DE MADUREZ DIGITAL DE LA EMPRESA EN EL"),
        ("EMPRESA CON POLITICAS DE SOSTENIBILIDAD:", "EMPRESA CON POLITICAS DE SOSTENIBILIDAD"),
        ("EMPRESA CON POLÍTICAS O PLANES DE TRANSFORMACIÓN DIGITAL:", "EMPRESA CON POLITICAS O PLANES DE TRANSFORMACION DIGITAL"),
        ("LA MÁXIMA RESPONSABLE DE LA EMPRESA O MÁS DEL 50% DEL EQUIPO DIRECTIVO ES MUJER:", "LA MAXIMA RESPONSABLE DE LA EMPRESA O MAS DEL 50% DEL EQUIPO DIRECTIVO"),
        # Campos exclusivos Directivos
        ("TITULACION",                "TITULACION"),
        ("NIF EMPRESA",               "NIF EMPRESA"),
        ("RELACIÓN CON LA EMPRESA",   "RELACION CON LA EMPRESA"),
        ("DEPARTAMENTO",              "DEPARTAMENTO"),
        ("PUESTO/CARGO",              "PUESTO/CARGO"),
        ("CANALES DE RELACION DE LA EMPRESA con clientes y proveedores.", "CANALES DE RELACION DE LA EMPRESA CON CLIENTES Y PROVEEDORES"),
        ("PROFESIONALES CON PERFIL TIC en la EMPRESA", "PROFESIONALES CON PERFIL TIC EN LA EMPRESA"),
        ("Describir motivación para cursar el programa", "DESCRIBIR MOTIVACION PARA CURSAR EL PROGRAMA"),
        # Campos exclusivos Agentes
        ("MOTIVACIÓN para cursar el programa", "MOTIVACION PARA CURSAR EL PROGRAMA"),
    ]

    for etiqueta_pdf, clave_norm in campos_linea:
        clave_dest = normalizar(clave_norm)
        # Si ya tenemos el campo de coordenadas y no estamos en modo corrompido, no sobreescribir
        if not pdf_corrompido and clave_dest in campos:
            continue
        val = extraer_linea(etiqueta_pdf, lineas)
        if val:
            val = re.sub(r'\s+', ' ', val).strip()
            campos[clave_dest] = val

    # Limpiar claves con ':' al final
    for k in list(campos.keys()):
        if k.endswith(':'):
            campos[k[:-1].strip()] = campos.pop(k)

    # ── Campos especiales con lógica propia ───────────────────────────────────

    # TIPO DE DOCUMENTO y Nº DE DOCUMENTO — en Directivos el PDF los tiene
    # en orden inverso visual (tipo en valor, nº en siguiente fila).
    # Asegurar que no están intercambiados.
    clave_tipo = normalizar("TIPO DE DOCUMENTO (NIF, NIE, DOCUMENTO ID, PASAPORTE)")
    clave_num  = normalizar("No DE DOCUMENTO")
    tipo_doc = campos.get(clave_tipo, "")
    num_doc  = campos.get(clave_num, "")
    TIPOS_VALIDOS = {"NIF", "NIE", "PASAPORTE", "DOCUMENTO ID", "PASSPORT"}
    # Si el valor de tipo_doc es claramente un número de documento y el de num_doc un tipo → intercambiar
    if tipo_doc and num_doc:
        tipo_parece_num = bool(re.search(r'\d{5,}', tipo_doc)) or not any(t in tipo_doc.upper() for t in TIPOS_VALIDOS)
        num_parece_tipo = any(t in num_doc.upper() for t in TIPOS_VALIDOS)
        if tipo_parece_num and num_parece_tipo:
            campos[clave_tipo], campos[clave_num] = num_doc, tipo_doc
    # Si tipo_doc contiene el tipo Y el número juntos (ej: "Y9722716D NIE"), separar
    if tipo_doc:
        for tipo in TIPOS_VALIDOS:
            if tipo in tipo_doc.upper():
                # Extraer el número: lo que no es el tipo
                numero_extraido = re.sub(re.escape(tipo), '', tipo_doc, flags=re.IGNORECASE).strip()
                if numero_extraido and not num_doc:
                    campos[clave_num]  = numero_extraido
                    campos[clave_tipo] = tipo
                break

    # RELACIÓN CON LA EMPRESA — valor partido alrededor de la etiqueta
    relacion = extraer_relacion_empresa_lineas(lineas)
    if relacion:
        campos[normalizar("RELACION CON LA EMPRESA")] = relacion

    # PORCENTAJE MUJERES
    clave_porc = normalizar("PORCENTAJE DE MUJERES CON RELACION LABORAL CON LA EMPRESA")
    if clave_porc not in campos:
        porcentaje = extraer_porcentaje_mujeres_lineas(lineas)
        if porcentaje:
            campos[clave_porc] = porcentaje

    # PERFIL DE LINKEDIN — el valor puede estar en la línea siguiente
    clave_linkedin = normalizar("PERFIL DE LINKEDIN")
    if clave_linkedin not in campos:
        linkedin = extraer_linea("Perfil de LinkedIn", lineas)
        if not linkedin:
            for i, l in enumerate(lineas):
                if l.strip() == "Perfil de LinkedIn" and i + 1 < len(lineas):
                    linkedin = lineas[i + 1].strip()
                    break
        if linkedin:
            campos[clave_linkedin] = linkedin

    # ACTIVIDAD DE LA EMPRESA — puede continuar en línea siguiente
    clave_act = normalizar("ACTIVIDAD DE LA EMPRESA (CODIGOS CNAE NIVEL LETRA)")
    if clave_act not in campos:
        actividad = extraer_linea(
            "ACTIVIDAD DE LA EMPRESA (códigos CNAE nivel letra)", lineas, multilinea=True
        )
        if actividad:
            campos[clave_act] = actividad

    # MOTIVACIÓN Directivos
    clave_motiv_dir = normalizar("DESCRIBIR MOTIVACION PARA CURSAR EL PROGRAMA")
    if clave_motiv_dir not in campos:
        motivacion_dir = extraer_linea("Describir motivación para cursar el programa", lineas, multilinea=True)
        if motivacion_dir:
            campos[clave_motiv_dir] = motivacion_dir

    # MOTIVACIÓN Agentes — puede estar en línea siguiente si el campo está vacío en el PDF
    clave_motiv_ag = normalizar("MOTIVACION PARA CURSAR EL PROGRAMA")
    if clave_motiv_ag not in campos:
        motivacion_ag = extraer_linea("MOTIVACIÓN para cursar el programa", lineas, multilinea=True)
        if not motivacion_ag:
            for i, l in enumerate(lineas):
                if re.search(r'MOTIVACI[OÓ]N\s+para cursar', l, re.IGNORECASE) and i + 1 < len(lineas):
                    sig = lineas[i + 1].strip()
                    if sig and not any(sig.upper().startswith(k.upper()) for k in ENCABEZADOS_STOP):
                        motivacion_ag = sig
                    break
        if motivacion_ag:
            campos[clave_motiv_ag] = motivacion_ag

    # Nº DE DOCUMENTO — en Aguilera (PDF firmado) pdfplumber puede no extraerlo;
    # intentar extracción por líneas si está vacío
    if clave_num not in campos or not campos.get(clave_num):
        num_extraido = extraer_linea("Nº de Documento", lineas)
        if num_extraido:
            campos[clave_num] = num_extraido

    # NIVEL DE MADUREZ DIGITAL — extracción por líneas como fallback
    clave_mad = normalizar("NIVEL DE MADUREZ DIGITAL DE LA EMPRESA EN EL")
    if clave_mad not in campos:
        madurez = extraer_linea(
            'NIVEL DE MADUREZ DIGITAL DE LA EMPRESA en el "Test de Diagnóstico Digital" de acelerapyme',
            lineas, multilinea=False
        )
        if madurez:
            campos[clave_mad] = madurez

    # Normalizar espacios en todos los valores
    for k in campos:
        if isinstance(campos[k], str):
            campos[k] = re.sub(r'\s+', ' ', campos[k]).strip()

    return campos


def postprocesar_campos(campos: dict) -> dict:
    """
    Limpia inconsistencias:
    - Campos con valores que son solo la etiqueta/descripción del campo.
    - Lista de opciones concatenadas con la respuesta.
    - Si desempleado, vacía datos de empresa.
    """
    # ── 1. Limpiar valores que son solo la etiqueta o texto descriptivo ────────
    BASURA = {"(empresa)", "empresa", ""}
    for k in list(campos.keys()):
        v = campos.get(k, "")
        if not isinstance(v, str):
            continue
        v_strip = v.strip()
        if v_strip.lower() in BASURA:
            del campos[k]
            continue
        # Si el valor empieza por '(empresa)' seguido de un valor real, limpiar el prefijo
        if v_strip.lower().startswith("(empresa)"):
            campos[k] = v_strip[len("(empresa)"):].strip()

    # ── 2. Limpiar campos de tipo lista-opciones ───────────────────────────────
    # Campos donde el PDF puede concatenar "lista_opciones RESPUESTA"
    claves_lista = [
        normalizar("ANTIGUEDAD DE LA EMPRESA"),
        normalizar("TAMANO EMPRESA"),
        normalizar("FACTURACION ULTIMO ANO"),
        normalizar("NIVEL DE MADUREZ DIGITAL DE LA EMPRESA EN EL"),
    ]
    for k in claves_lista:
        if k in campos and isinstance(campos[k], str):
            campos[k] = _limpiar_opcion_lista(campos[k])

    # ── 3. Limpiar AMBITO RURAL (a veces trae la descripción o SI/NO doble) ───
    clave_ambito = normalizar("AMBITO RURAL")
    if clave_ambito in campos:
        v = campos[clave_ambito]
        if isinstance(v, str):
            # Extraer el último SI/NO
            matches = list(re.finditer(r'\b(SI|NO)\b', v, re.IGNORECASE))
            if matches:
                campos[clave_ambito] = matches[-1].group(1).upper()
            elif "municipio" in v.lower():
                del campos[clave_ambito]

    # ── 4. Limpiar RESIDE que pypdf parte en dos líneas ("...5.000: SI / NO NO") ──
    clave_reside = normalizar("RESIDE EN UNA LOCALIDAD CON UN NUMERO DE HABITANTES INFERIOR A 5")
    if clave_reside in campos:
        v = campos[clave_reside]
        if isinstance(v, str):
            matches = list(re.finditer(r'\b(SI|NO)\b', v, re.IGNORECASE))
            if matches:
                campos[clave_reside] = matches[-1].group(1).upper()

    # ── 5. Normalizar espacios (colapsar múltiples espacios) ──────────────────
    for k in list(campos.keys()):
        if isinstance(campos[k], str):
            campos[k] = re.sub(r'\s+', ' ', campos[k]).strip()

    # Campos específicos donde pdfplumber fragmenta palabras — limpiar espacios internos
    # (ej: "Si n experi enci a" → "Sin experiencia", "M - A ctividade s" → "M - Actividades")
    CAMPOS_FRAGMENTADOS = [
        normalizar("EXPERIENCIA LABORAL EN PUESTOS DE DIGITALIZACION"),
        normalizar("ACTIVIDAD DE LA EMPRESA (CODIGOS CNAE NIVEL LETRA)"),
        normalizar("NOMBRE EMPRESA (RAZON SOCIAL)"),
    ]
    for k in CAMPOS_FRAGMENTADOS:
        if k in campos and isinstance(campos[k], str):
            # Eliminar espacios sueltos dentro de tokens: " c ti vi da de s" → "actividades"
            # Solo si los segmentos son muy cortos (1-2 chars) — señal de fragmentación
            palabras = campos[k].split()
            resultado = []
            i = 0
            while i < len(palabras):
                p = palabras[i]
                # Agrupar fragmentos cortos consecutivos
                while i + 1 < len(palabras) and len(palabras[i+1]) <= 2 and len(p) <= 4:
                    i += 1
                    p += palabras[i]
                resultado.append(p)
                i += 1
            campos[k] = ' '.join(resultado)

    # ── 6. Limpiar datos de empresa si desempleado ─────────────────────────────
    clave_sit = normalizar("SITUACION LABORAL ACTUAL")
    situacion = buscar_valor_en_campos(campos, clave_sit) or ""

    def _es_desempleo(texto: str) -> bool:
        t = str(texto).strip().lower()
        # Captura variantes: desempleado, desempleada, desempleado/a, en desempleo, etc.
        return bool(re.search(r'\bdesemplead[oa]?\b|\bdesemplead[oa]?/a\b|\bdesempleo\b|\bsin\s+empleo\b|\ben\s+paro\b', t))

    if _es_desempleo(situacion):
        claves_empresa_raw = [
            "NOMBRE EMPRESA (RAZON SOCIAL)", "NIF", "DEPARTAMENTO (EMPRESA)",
            "PUESTO/CARGO (EMPRESA)", "ACTIVIDAD DE LA EMPRESA (CODIGOS CNAE NIVEL LETRA)",
            "TAMANO EMPRESA", "DIRECCION (EMPRESA)", "CIUDAD (EMPRESA)",
            "CODIGO POSTAL (EMPRESA)", "CCAA (EMPRESA)", "PROVINCIA (EMPRESA)",
            "TELEFONO (EMPRESA)", "PAGINA WEB (EMPRESA)", "ANTIGUEDAD DE LA EMPRESA",
            "FACTURACION ULTIMO ANO", "AMBITO RURAL",
            "NIVEL DE MADUREZ DIGITAL DE LA EMPRESA EN EL",
            "EMPRESA CON POLITICAS DE SOSTENIBILIDAD",
            "EMPRESA CON POLITICAS O PLANES DE TRANSFORMACION DIGITAL",
            "LA MAXIMA RESPONSABLE DE LA EMPRESA O MAS DEL 50% DEL EQUIPO DIRECTIVO",
            "PORCENTAJE DE MUJERES CON RELACION LABORAL CON LA EMPRESA",
            # Variantes frecuentes / campos directivos
            "NIF EMPRESA", "DEPARTAMENTO", "PUESTO/CARGO", "RELACION CON LA EMPRESA",
            "CANALES DE RELACION DE LA EMPRESA CON CLIENTES Y PROVEEDORES",
            "PROFESIONALES CON PERFIL TIC EN LA EMPRESA",
        ]
        claves_empresa = {normalizar(c) for c in claves_empresa_raw}
        prefijos_empresa = {normalizar(p) for p in [
            "DIRECCION (EMPRESA", "CIUDAD (EMPRESA", "CODIGO POSTAL (EMPRESA",
            "CCAA (EMPRESA", "PROVINCIA (EMPRESA", "TELEFONO (EMPRESA",
            "PAGINA WEB (EMPRESA", "DEPARTAMENTO (EMPRESA", "PUESTO/CARGO (EMPRESA",
            "NOMBRE EMPRESA", "NIF EMPRESA", "ACTIVIDAD DE LA EMPRESA",
        ]}
        for k in list(campos.keys()):
            if k in claves_empresa or any(k.startswith(p) for p in prefijos_empresa):
                del campos[k]
    return campos


def buscar_valor_en_campos(campos: dict, clave_mapeo: str) -> str | None:
    """
    Busca un valor en el diccionario de campos.
    Primero exacto; luego la clave del PDF empieza por la clave del mapeo;
    finalmente la clave del mapeo empieza por la del PDF (solo si len > 8).
    Se evita devolver valores de claves demasiado cortas y ambiguas.
    """
    if clave_mapeo in campos:
        return campos[clave_mapeo]
    for clave_pdf, valor in campos.items():
        if clave_pdf.startswith(clave_mapeo) and len(clave_mapeo) > 5:
            return valor
    for clave_pdf, valor in campos.items():
        if clave_mapeo.startswith(clave_pdf) and len(clave_pdf) > 8:
            return valor
    return None


# ──────────────────────────────────────────────────────────────────────────────
# ESCRITURA EN EL FORMULARIO EXCEL
# ──────────────────────────────────────────────────────────────────────────────

def rellenar_formulario(campos: dict, ruta_salida: str, tipo: str):
    """
    Decodifica la plantilla embebida, la copia a ruta_salida y la rellena.
    """
    mapeo         = MAPEO_POR_TIPO[tipo]
    hoja          = HOJA_POR_TIPO[tipo]
    filas_si_no   = FILAS_SI_NO_POR_TIPO[tipo]
    declaraciones = DECLARACIONES_POR_TIPO[tipo]
    fila_fact     = 43 if tipo == TIPO_DIRECTIVOS else 48
    fila_madurez  = 45 if tipo == TIPO_DIRECTIVOS else 50
    fila_antiguedad = 42 if tipo == TIPO_DIRECTIVOS else 47
    fila_tamano     = 34 if tipo == TIPO_DIRECTIVOS else 39

    # Extraer plantilla embebida a fichero temporal y copiar a destino
    ruta_tmp = obtener_ruta_plantilla(tipo)
    try:
        shutil.copy2(ruta_tmp, ruta_salida)
    finally:
        os.unlink(ruta_tmp)

    wb = openpyxl.load_workbook(ruta_salida, keep_vba=True)
    ws = wb[hoja]

    for clave_mapeo, fila in mapeo.items():
        valor = buscar_valor_en_campos(campos, normalizar(clave_mapeo))
        if valor is not None:
            if fila == 12:
                valor = parsear_fecha(valor)
            elif fila in filas_si_no:
                valor = _extraer_si_no(valor)
            elif fila == fila_fact:
                valor = _limpiar_facturacion(valor)
            elif fila == fila_madurez:
                valor = _limpiar_madurez(valor)
            elif fila in (fila_antiguedad, fila_tamano):
                valor = _limpiar_opcion_lista(valor)
            ws.cell(fila, 2).value = valor

    for fila_decl in declaraciones:
        ws.cell(fila_decl, 2).value = "Acepto"

    wb.save(ruta_salida)


# ──────────────────────────────────────────────────────────────────────────────
# INTERFAZ TKINTER  (ventana única para toda la sesión)
# ──────────────────────────────────────────────────────────────────────────────

def pedir_limite_formularios(root) -> int | None:
    resultado = {"valor": None}

    ventana = tk.Toplevel(root)
    ventana.title("Límite de formularios")
    ventana.resizable(False, False)
    ventana.grab_set()

    ventana.update_idletasks()
    ancho, alto = 340, 160
    x = (ventana.winfo_screenwidth() - ancho) // 2
    y = (ventana.winfo_screenheight() - alto) // 2
    ventana.geometry(f"{ancho}x{alto}+{x}+{y}")

    tk.Label(
        ventana,
        text="¿Cuántos formularios como máximo\nquieres procesar en esta ejecución?",
        font=("Segoe UI", 10),
        pady=12,
    ).pack()

    frame = tk.Frame(ventana)
    frame.pack()
    tk.Label(frame, text="Límite:", font=("Segoe UI", 10)).pack(side="left", padx=6)
    entrada = tk.Entry(frame, width=8, font=("Segoe UI", 11), justify="center")
    entrada.insert(0, "10")
    entrada.pack(side="left")
    entrada.focus_set()
    entrada.select_range(0, tk.END)

    def confirmar(event=None):
        texto = entrada.get().strip()
        if not texto.isdigit() or int(texto) < 1:
            messagebox.showwarning("Valor inválido", "Introduce un número entero mayor que 0.", parent=ventana)
            return
        resultado["valor"] = int(texto)
        ventana.destroy()

    def cancelar():
        ventana.destroy()

    frame_btn = tk.Frame(ventana, pady=10)
    frame_btn.pack()
    tk.Button(frame_btn, text="Aceptar",    width=10, command=confirmar).pack(side="left", padx=6)
    tk.Button(frame_btn, text="Sin límite", width=10, command=lambda: [resultado.update({"valor": 0}), ventana.destroy()]).pack(side="left", padx=6)
    tk.Button(frame_btn, text="Cancelar",   width=10, command=cancelar).pack(side="left", padx=6)

    entrada.bind("<Return>", confirmar)
    ventana.wait_window()
    return resultado["valor"]


def seleccionar_archivos(root):
    """Recibe el root ya creado y devuelve (rutas_pdf, carpeta_salida, limite)."""
    limite = pedir_limite_formularios(root)
    if limite is None:
        messagebox.showinfo("Cancelado", "Operación cancelada por el usuario.", parent=root)
        return None, None, None

    messagebox.showinfo("Paso 1/2", "Selecciona uno o varios PDF de formularios", parent=root)
    rutas_pdf = filedialog.askopenfilenames(
        title="Formularios PDF",
        filetypes=[("PDF", "*.pdf"), ("Todos", "*.*")],
        parent=root,
    )
    if not rutas_pdf:
        messagebox.showerror("Cancelado", "No se seleccionaron PDFs.", parent=root)
        return None, None, None

    messagebox.showinfo("Paso 2/2", "Selecciona la carpeta de SALIDA donde se guardarán los archivos", parent=root)
    carpeta_salida = filedialog.askdirectory(title="Carpeta de salida", parent=root)
    if not carpeta_salida:
        messagebox.showerror("Cancelado", "No se seleccionó carpeta de salida.", parent=root)
        return None, None, None

    return list(rutas_pdf), carpeta_salida, limite


def guardar_log_pdfs_ilegibles(carpeta_salida: str, pdfs_ilegibles: list[tuple[str, str, str]]) -> str | None:
    if not pdfs_ilegibles:
        return None
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta_log = os.path.join(carpeta_salida, f"log_pdfs_ilegibles_{timestamp}.txt")
    with open(ruta_log, "w", encoding="utf-8") as f:
        f.write("LOG DE PDF ILEGIBLES\n")
        f.write(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total: {len(pdfs_ilegibles)}\n\n")
        for idx, (pdf, tipo_error, detalle) in enumerate(pdfs_ilegibles, start=1):
            f.write(f"{idx}. PDF: {pdf}\n")
            f.write(f"   Error: {tipo_error}\n")
            f.write(f"   Detalle: {detalle}\n\n")
    return ruta_log


def imprimir_seguro(texto: str):
    """Imprime en consola evitando errores de codificación en terminales Windows."""
    try:
        print(texto)
    except UnicodeEncodeError:
        # Fallback: eliminar caracteres no representables en la codificación actual
        encoding = getattr(getattr(__import__('sys'), 'stdout', None), 'encoding', None) or 'utf-8'
        safe = texto.encode(encoding, errors='replace').decode(encoding, errors='replace')
        print(safe)


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    # Una sola instancia de Tk para toda la sesión
    root = tk.Tk()
    root.withdraw()

    rutas_pdf, carpeta_salida, limite = seleccionar_archivos(root)
    if rutas_pdf is None:
        root.destroy()
        return

    total_seleccionados = len(rutas_pdf)
    omitidos = 0
    if limite and limite > 0 and total_seleccionados > limite:
        omitidos = total_seleccionados - limite
        rutas_pdf = rutas_pdf[:limite]
        imprimir_seguro(f"AVISO: Límite de {limite} formularios: se omiten {omitidos} PDF(s).")

    errores    = []
    errores_rellenado = []
    procesados = 0
    pdfs_ilegibles = []

    for i, ruta_pdf in enumerate(rutas_pdf, start=1):
        nombre_pdf = os.path.basename(ruta_pdf)
        try:
            lineas = leer_lineas_pdf(ruta_pdf)
            if not lineas:
                raise ValueError("Sin texto extraíble")
        except Exception as e:
            pdfs_ilegibles.append((nombre_pdf, type(e).__name__, str(e)))
            errores.append(f"{nombre_pdf}: PDF ilegible ({e})")
            imprimir_seguro(f"  ERROR [{i}/{len(rutas_pdf)}] {nombre_pdf}: PDF ilegible ({e})")
            continue

        try:
            # 2. Detectar tipo usando las líneas ya leídas
            tipo = detectar_tipo_pdf(lineas)
            imprimir_seguro(f"Procesando [{i}/{len(rutas_pdf)}]: {nombre_pdf} -> tipo: {tipo}")

            # 3. Extraer y postprocesar campos
            campos = leer_campos_pdf(ruta_pdf, lineas)
            campos = postprocesar_campos(campos)
            imprimir_seguro(f"  Campos extraidos: {len(campos)}")

            # 4. Generar nombre de salida
            apellido = buscar_valor_en_campos(campos, normalizar("PRIMER APELLIDO")) or "alumno"
            apellido = re.sub(r'[^\w]', '_', str(apellido))
            nombre_salida = f"Formulario_{apellido}_{i}.xlsm"
            ruta_salida   = os.path.join(carpeta_salida, nombre_salida)

            # 5. Rellenar y guardar (plantilla embebida, sin diálogo)
            rellenar_formulario(campos, ruta_salida, tipo)
            imprimir_seguro(f"  Guardado: {nombre_salida}")

            procesados += 1

        except Exception as e:
            errores_rellenado.append((nombre_pdf, type(e).__name__, str(e)))
            errores.append(f"{nombre_pdf}: {e}")
            imprimir_seguro(f"  ERROR [{i}/{len(rutas_pdf)}] {nombre_pdf}: {e}")

    ruta_log_ilegibles = None
    try:
        ruta_log_ilegibles = guardar_log_pdfs_ilegibles(carpeta_salida, pdfs_ilegibles)
    except Exception as e:
        errores.append(f"No se pudo guardar el log de PDF ilegibles: {e}")
        imprimir_seguro(f"  ERROR: no se pudo guardar log de PDF ilegibles ({e})")

    # Resumen final
    total_ilegibles = len(pdfs_ilegibles)
    total_rellenado = len(errores_rellenado)
    resumen = (
        f"Procesados OK: {procesados} de {total_seleccionados} PDF(s)"
        f"\nIlegibles: {total_ilegibles}"
        f"\nErrores de rellenado: {total_rellenado}"
    )
    if omitidos:
        resumen += f"\n\nAVISO: {omitidos} PDF(s) omitidos por limite de {limite} formularios."
    if ruta_log_ilegibles:
        resumen += f"\n\nLog de PDF ilegibles: {ruta_log_ilegibles}"
    if errores_rellenado:
        resumen += f"\n\nDetalle errores de rellenado ({total_rellenado}):\n"
        resumen += "\n".join(f"{pdf}: {tipo} ({detalle})" for pdf, tipo, detalle in errores_rellenado)
    elif errores:
        resumen += f"\n\nErrores ({len(errores)}):\n" + "\n".join(errores)

    try:
        messagebox.showinfo("Proceso completado", resumen, parent=root)
    except Exception:
        # Si no hay UI disponible, al menos mostrar por consola
        pass
    root.destroy()
    imprimir_seguro("\n" + resumen)


if __name__ == "__main__":
    main()
