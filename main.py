"""
╔════════════════════════════════════════════════════════════╗
║   PDF Formulario Agentes / Directivos  →  Excel            ║
║   Grupo ATU © 2026                                         ║
║   Hecho por RaulRDA.com, Pablo Álvarez y Pelayo Fernández  ║
╚════════════════════════════════════════════════════════════╝
"""

import re
import io
import shutil
import sys
import traceback
import webbrowser
import requests
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext

import pdfplumber
import openpyxl
from pypdf import PdfReader
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string


# ── RUTA BASE ────────────────────────────────────────────────
def get_base_dir() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys._MEIPASS)
    return Path(__file__).parent

BASE_DIR = get_base_dir()

# ── VERSIÓN ──────────────────────────────────────────────────
VERSION    = "1.1.0"   # versión actual del programa
UPDATE_URL = "https://github.com/RaulRDA/convierte-tu/releases/latest"   # URL pública de releases

# ── PALETA ATU ───────────────────────────────────────────────
ATU_NAVY    = "#1C2B4A"
ATU_ORANGE  = "#E8721C"
ATU_LIGHT   = "#F5F7FA"
ATU_WHITE   = "#FFFFFF"
ATU_TEXT    = "#2D3748"
ATU_MID     = "#4A6FA5"
ATU_SUCCESS = "#2E7D32"
ATU_ERROR   = "#C62828"
ATU_WARN    = "#E65100"
ATU_BORDER  = "#CBD5E0"

# ── TIPOS ─────────────────────────────────────────────────────
TIPO_DIRECTIVOS = "Directivos"
TIPO_AGENTES_1  = "Agentes (1)"
TIPO_AGENTES_2  = "Agentes (2)"

# ── PLANTILLAS ───────────────────────────────────────────────
PLANTILLAS = {
    TIPO_DIRECTIVOS: ("_plantillas/formulario_directivos.xlsm", "Formulario Directivos", 50),
    TIPO_AGENTES_1:  ("_plantillas/formulario_agentes1.xlsm",   "formulario agentes",    40),
    TIPO_AGENTES_2:  ("_plantillas/formulario_agentes2.xlsm",   "formulario agentes",    40),
}

# Separador horizontal entre etiqueta y valor en el PDF
SEP_X = 350

# ── MAPEOS ───────────────────────────────────────────────────
MAPEO_DIRECTIVOS = {
    "PRIMER APELLIDO":            "B6",
    "SEGUNDO APELLIDO":           "B7",
    "NOMBRE":                     "B8",
    "TIPO DE DOCUMENTO":          "B9",
    "NUM DE DOCUMENTO":           "B10",
    "SEXO":                       "B11",
    "FECHA DE NACIMIENTO":        "B12",
    "DIRECCION":                  "B13",
    "CIUDAD":                     "B14",
    "CODIGO POSTAL":              "B15",
    "CCAA":                       "B16",
    "PROVINCIA":                  "B17",
    "TELEFONO":                   "B18",
    "EMAIL":                      "B19",
    "RESIDE LOCALIDAD PEQUENA":   "B20",
    "PERSONA CON DISCAPACIDAD":   "B21",
    "NIVEL DE ESTUDIOS":          "B22",
    "TITULACION":                 "B23",
    "NOMBRE EMPRESA":             "B26",
    "RELACION CON LA EMPRESA":    "B27",
    "DEPARTAMENTO EMPRESA":       "B28",
    "PUESTO EMPRESA":             "B29",
    "NIF EMPRESA":                "B32",
    "ACTIVIDAD EMPRESA":          "B33",
    "TAMANO EMPRESA":             "B34",
    "DIRECCION EMPRESA":          "B35",
    "CIUDAD EMPRESA":             "B36",
    "CODIGO POSTAL EMPRESA":      "B37",
    "CCAA EMPRESA":               "B38",
    "PROVINCIA EMPRESA":          "B39",
    "TELEFONO EMPRESA":           "B40",
    "WEB EMPRESA":                "B41",
    "ANTIGUEDAD EMPRESA":         "B42",
    "FACTURACION EMPRESA":        "B43",
    "AMBITO RURAL EMPRESA":       "B44",
    "MADUREZ DIGITAL EMPRESA":    "B45",
    "CANALES RELACION EMPRESA":   "B46",
    "PERFIL TIC EMPRESA":         "B47",
    "SOSTENIBILIDAD EMPRESA":     "B48",
    "PLAN DIGITAL EMPRESA":       "B49",
    "MUJER DIRECTIVA EMPRESA":    "B50",
    "PORCENTAJE MUJERES EMPRESA": "B51",
    "MOTIVACION":                 "B54",
    "ACEPTO CONDICIONADO":        "B85",
    "DECLARO PYME":               "B86",
    "DECLARO REALIDAD":           "B87",
    "DECLARO NO RECIBIDA":        "B88",
    "DECLARO CONFLICTO":          "B89",
    "AUTORIZO DATOS":             "B90",
    "ACEPTO DISCAPACIDAD":        "B91",
}

MAPEO_AGENTES2 = {
    "PRIMER APELLIDO":             "B6",
    "SEGUNDO APELLIDO":            "B7",
    "NOMBRE":                      "B8",
    "TIPO DE DOCUMENTO":           "B9",
    "NUM DE DOCUMENTO":            "B10",
    "SEXO":                        "B11",
    "FECHA DE NACIMIENTO":         "B12",
    "DIRECCION":                   "B13",
    "CIUDAD":                      "B14",
    "CODIGO POSTAL":               "B15",
    "CCAA":                        "B16",
    "PROVINCIA":                   "B17",
    "TELEFONO":                    "B18",
    "EMAIL":                       "B19",
    "RESIDE LOCALIDAD PEQUENA":    "B20",
    "PERSONA CON DISCAPACIDAD":    "B21",
    "PERFIL LINKEDIN":             "B22",
    "NIVEL DE ESTUDIOS":           "B25",
    "FORMACION DIGITALIZACION":    "B26",
    "FORMACION GESTION PROYECTOS": "B27",
    "ANOS EXPERIENCIA LABORAL":    "B30",
    "EXPERIENCIA DIGITALIZACION":  "B31",
    "SITUACION LABORAL":           "B32",
    "NOMBRE EMPRESA":              "B34",
    "NIF EMPRESA":                 "B35",
    "DEPARTAMENTO EMPRESA":        "B36",
    "PUESTO EMPRESA":              "B37",
    "ACTIVIDAD EMPRESA":           "B38",
    "TAMANO EMPRESA":              "B39",
    "DIRECCION EMPRESA":           "B40",
    "CIUDAD EMPRESA":              "B41",
    "CODIGO POSTAL EMPRESA":       "B42",
    "CCAA EMPRESA":                "B43",
    "PROVINCIA EMPRESA":           "B44",
    "TELEFONO EMPRESA":            "B45",
    "WEB EMPRESA":                 "B46",
    "ANTIGUEDAD EMPRESA":          "B47",
    "FACTURACION EMPRESA":         "B48",
    "AMBITO RURAL EMPRESA":        "B49",
    "MADUREZ DIGITAL EMPRESA":     "B50",
    "SOSTENIBILIDAD EMPRESA":      "B51",
    "PLAN DIGITAL EMPRESA":        "B52",
    "MUJER DIRECTIVA EMPRESA":     "B53",
    "PORCENTAJE MUJERES EMPRESA":  "B54",
    "MOTIVACION":                  "B57",
    "ACEPTO CONDICIONADO":         "B79",
    "DECLARO PYME":                "B80",
    "DECLARO REALIDAD":            "B81",
    "DECLARO NO RECIBIDA":         "B82",
    "DECLARO CONFLICTO":           "B83",
    "AUTORIZO DATOS":              "B84",
    "ACEPTO DISCAPACIDAD":         "B85",
}

MAPEO_AGENTES1 = dict(MAPEO_AGENTES2)
MAPEO_AGENTES1.pop("DECLARO PYME")
MAPEO_AGENTES1["ACEPTO CONDICIONADO"] = "B80"

MAPEO_POR_TIPO = {
    TIPO_DIRECTIVOS: MAPEO_DIRECTIVOS,
    TIPO_AGENTES_1:  MAPEO_AGENTES1,
    TIPO_AGENTES_2:  MAPEO_AGENTES2,
}

# Campos booleanos SI/NO
CAMPOS_SI_NO = {
    "RESIDE LOCALIDAD PEQUENA", "PERSONA CON DISCAPACIDAD",
    "AMBITO RURAL EMPRESA", "SOSTENIBILIDAD EMPRESA",
    "PLAN DIGITAL EMPRESA", "MUJER DIRECTIVA EMPRESA",
}

# Campos de empresa que se vacían si el alumno está desempleado
CAMPOS_EMPRESA = [
    "NOMBRE EMPRESA", "NIF EMPRESA", "DEPARTAMENTO EMPRESA",
    "PUESTO EMPRESA", "ACTIVIDAD EMPRESA", "TAMANO EMPRESA",
    "DIRECCION EMPRESA", "CIUDAD EMPRESA", "CODIGO POSTAL EMPRESA",
    "CCAA EMPRESA", "PROVINCIA EMPRESA", "TELEFONO EMPRESA",
    "WEB EMPRESA", "ANTIGUEDAD EMPRESA", "FACTURACION EMPRESA",
    "AMBITO RURAL EMPRESA", "MADUREZ DIGITAL EMPRESA",
    "SOSTENIBILIDAD EMPRESA", "PLAN DIGITAL EMPRESA",
    "MUJER DIRECTIVA EMPRESA", "PORCENTAJE MUJERES EMPRESA",
    "CANALES RELACION EMPRESA", "PERFIL TIC EMPRESA",
    "RELACION CON LA EMPRESA",
]

# Declaraciones que siempre se marcan "Acepto"
DECLARACIONES_COMUNES = [
    "ACEPTO CONDICIONADO", "DECLARO REALIDAD", "DECLARO NO RECIBIDA",
    "DECLARO CONFLICTO", "AUTORIZO DATOS", "ACEPTO DISCAPACIDAD",
]

# Campos reales que validan que la extracción tuvo éxito
CAMPOS_REALES = [
    "PRIMER APELLIDO", "SEGUNDO APELLIDO", "NOMBRE",
    "EMAIL", "TELEFONO", "NIF EMPRESA", "NOMBRE EMPRESA", "DIRECCION",
]

# Palabras de inicio de campo (para cortar extracción multilinea)
ENCABEZADOS_STOP = {
    "DATOS", "NOMBRE", "NIF", "RELACION", "DEPARTAMENTO", "PUESTO",
    "ACTIVIDAD", "TAMANO", "DIRECCION", "CIUDAD", "CODIGO", "CCAA",
    "PROVINCIA", "TELEFONO", "PAGINA", "ANTIGUEDAD", "FACTURACION",
    "AMBITO", "NIVEL", "CANALES", "PROFESIONALES", "EMPRESA",
    "LA MAXIMA", "PORCENTAJE", "SEGUNDO", "TIPO", "NUM", "SEXO",
    "FECHA", "EMAIL", "RESIDE", "PERSONA", "TITULACION",
    "PRIMER", "ACEPTO", "AUTORIZO", "ANOS", "EXPERIENCIA",
    "SITUACION", "MOTIVACION", "FORMACION", "PERFIL",
}


# ── UTILIDADES ───────────────────────────────────────────────
def _set_icon(ventana):
    ico = BASE_DIR / "icono.ico"
    if ico.exists():
        try:
            ventana.iconbitmap(str(ico))
        except Exception:
            pass

def limpiar_texto(texto: str) -> str:
    if not texto or not isinstance(texto, str):
        return ""
    return re.sub(r'\s+', ' ', texto.replace('\n', ' ').replace('\r', ' ')).strip()

def normalizar(texto):
    """Mayúsculas, sin tildes, espacios colapsados. Clave interna."""
    if not texto:
        return ""
    trans = str.maketrans(
        "áéíóúàèìòùâêîôûäëïöüÁÉÍÓÚÀÈÌÒÙÂÊÎÔÛÄËÏÖÜñÑ",
        "aeiouaeiouaeiouaeiouAEIOUAEIOUAEIOUAEIOUnN"
    )
    resultado = texto.upper().translate(trans).replace("º", "")
    return " ".join(resultado.split())


# ── POST-PROCESADO DE VALORES ────────────────────────────────
def extraer_si_no(valor):
    matches = list(re.finditer(r"\b(SI|NO)\b", str(valor), re.IGNORECASE))
    return matches[-1].group(1).upper() if matches else valor

def limpiar_opcion_lista(valor):
    valor = str(valor).strip()
    if "/" not in valor:
        return valor
    valor_n = valor.replace("\u2013", "-").replace("\u2014", "-")
    opciones = [p.strip() for p in valor_n.split("/")]
    ultima = opciones[-1]
    for opcion in opciones[:-1]:
        if ultima.startswith(opcion):
            resto = ultima[len(opcion):].strip()
            if resto:
                return resto
    for opcion in opciones:
        if opcion.strip() and ultima.endswith(opcion.strip()):
            return opcion.strip()
    return ultima

def limpiar_facturacion(valor):
    valor = str(valor).strip()
    m = re.search(r"([+\d][^/]*)\s*$", valor)
    if m:
        return m.group(1).strip()
    partes = valor.split()
    for i in range(len(partes) - 1, -1, -1):
        if re.match(r"^[\d+]", partes[i]):
            return " ".join(partes[i:])
    return valor

def limpiar_madurez(valor):
    niveles = ("bajo", "medio", "alto", "muy", "basico")
    partes = str(valor).split()
    for i, p in enumerate(partes):
        if p.lower() in niveles:
            return " ".join(partes[i:])
    return valor

def parsear_fecha(texto):
    meses = {
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
        "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
        "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    }
    texto_s = str(texto).strip()
    m = re.match(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})", texto_s)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", texto_s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    tl = texto_s.lower()
    for nombre, num in meses.items():
        if nombre in tl:
            partes = re.findall(r"\d+", tl)
            if len(partes) >= 2:
                try:
                    return datetime(int(partes[-1]), num, int(partes[0]))
                except ValueError:
                    pass
    return texto_s

def corregir_tipo_num_documento(datos):
    TIPOS_VALIDOS = {"NIF", "NIE", "PASAPORTE", "DOCUMENTO ID", "PASSPORT"}
    tipo_doc = datos.get("TIPO DE DOCUMENTO", "")
    num_doc  = datos.get("NUM DE DOCUMENTO", "")

    if tipo_doc:
        for tipo in TIPOS_VALIDOS:
            if tipo in tipo_doc.upper():
                numero_extraido = re.sub(
                    re.escape(tipo), "", tipo_doc, flags=re.IGNORECASE
                ).strip()
                if numero_extraido and not num_doc:
                    datos["NUM DE DOCUMENTO"]  = numero_extraido
                    datos["TIPO DE DOCUMENTO"] = tipo
                break

    tipo_doc = datos.get("TIPO DE DOCUMENTO", "")
    num_doc  = datos.get("NUM DE DOCUMENTO", "")

    if tipo_doc and num_doc:
        tipo_parece_num = (
            bool(re.search(r"\d{5,}", tipo_doc))
            or not any(t in tipo_doc.upper() for t in TIPOS_VALIDOS)
        )
        num_parece_tipo = any(t in num_doc.upper() for t in TIPOS_VALIDOS)
        if tipo_parece_num and num_parece_tipo:
            datos["TIPO DE DOCUMENTO"], datos["NUM DE DOCUMENTO"] = num_doc, tipo_doc

    return datos

def postprocesar_campos(datos, tipo):
    # 1. Limpiar basura
    for k in list(datos.keys()):
        v = datos.get(k, "")
        if isinstance(v, str):
            v = re.sub(r"\s+", " ", v).strip()
            if v.lower() in {"", "(empresa)", "empresa"}:
                del datos[k]
                continue
            if v.lower().startswith("(empresa)"):
                v = v[len("(empresa)"):].strip()
            datos[k] = v

    # 2. Listas de opciones
    for campo in ("ANTIGUEDAD EMPRESA", "TAMANO EMPRESA", "MADUREZ DIGITAL EMPRESA"):
        if campo in datos:
            datos[campo] = limpiar_opcion_lista(datos[campo])

    # 3. Facturación
    if "FACTURACION EMPRESA" in datos:
        datos["FACTURACION EMPRESA"] = limpiar_facturacion(datos["FACTURACION EMPRESA"])

    # 4. Madurez digital
    if "MADUREZ DIGITAL EMPRESA" in datos:
        datos["MADUREZ DIGITAL EMPRESA"] = limpiar_madurez(datos["MADUREZ DIGITAL EMPRESA"])

    # 5. SI/NO en campos booleanos
    for campo in CAMPOS_SI_NO:
        if campo in datos:
            datos[campo] = extraer_si_no(datos[campo])

    # 6. Ámbito rural puede traer texto largo
    if "AMBITO RURAL EMPRESA" in datos:
        v = datos["AMBITO RURAL EMPRESA"]
        matches = list(re.finditer(r"\b(SI|NO)\b", v, re.IGNORECASE))
        if matches:
            datos["AMBITO RURAL EMPRESA"] = matches[-1].group(1).upper()
        elif "municipio" in v.lower():
            del datos["AMBITO RURAL EMPRESA"]

    # 7. Fecha de nacimiento
    if "FECHA DE NACIMIENTO" in datos:
        datos["FECHA DE NACIMIENTO"] = parsear_fecha(datos["FECHA DE NACIMIENTO"])

    # 8. Vaciar empresa si desempleado
    situacion = datos.get("SITUACION LABORAL", "")
    es_desempleo = bool(re.search(
        r"\bdesemplead[oa]?\b|\bdesempleo\b|\bsin\s+empleo\b|\ben\s+paro\b",
        str(situacion).lower()
    ))
    if es_desempleo:
        for campo in CAMPOS_EMPRESA:
            datos.pop(campo, None)

    # 9. Corrección tipo/num documento
    datos = corregir_tipo_num_documento(datos)

    return datos


# ── LECTURA DE LÍNEAS DEL PDF ────────────────────────────────
def leer_lineas_pdf(ruta_pdf):
    errores = []
    try:
        with open(ruta_pdf, "rb") as fh:
            reader = PdfReader(fh)
            trozos = []
            for page in reader.pages:
                try:
                    trozos.append(page.extract_text() or "")
                except Exception as e:
                    errores.append(str(e))
        lineas = [l.strip() for l in "\n".join(trozos).splitlines() if l.strip()]
        if lineas:
            return lineas
        errores.append("pypdf: sin texto")
    except Exception as e:
        errores.append("pypdf: " + str(e))

    try:
        with pdfplumber.open(ruta_pdf) as pdf:
            texto = "\n".join((p.extract_text() or "") for p in pdf.pages)
        lineas = [l.strip() for l in texto.splitlines() if l.strip()]
        if lineas:
            return lineas
        errores.append("pdfplumber: sin texto")
    except Exception as e:
        errores.append("pdfplumber: " + str(e))

    raise ValueError("PDF ilegible. " + " | ".join(errores))


# ── ESTRATEGIA 1: TABLAS (pdfplumber) ────────────────────────
def _mapear_campo_tabla(etiq, valor, datos, tipo):
    e  = etiq.upper()
    en = normalizar(etiq)

    if "PRIMER APELLIDO" in e:
        datos["PRIMER APELLIDO"] = valor
    elif "SEGUNDO APELLIDO" in e:
        datos["SEGUNDO APELLIDO"] = valor
    elif "NOMBRE" in e and "EMPRESA" not in e and "RAZON" not in e and "SOCIAL" not in e:
        datos["NOMBRE"] = valor
    elif "TIPO DE DOCUMENTO" in e:
        datos["TIPO DE DOCUMENTO"] = valor
    elif ("NUM" in en or "Nº" in e or "N°" in e) and "DOCUMENTO" in e:
        datos["NUM DE DOCUMENTO"] = valor
    elif "SEXO" in e:
        datos["SEXO"] = valor
    elif "FECHA DE NACIMIENTO" in e:
        datos["FECHA DE NACIMIENTO"] = valor
    elif "DIRECCION" in en and "EMPRESA" not in en:
        datos["DIRECCION"] = valor
    elif "CIUDAD" in e and "EMPRESA" not in e:
        datos["CIUDAD"] = valor
    elif "CODIGO POSTAL" in e and "EMPRESA" not in e:
        datos["CODIGO POSTAL"] = valor
    elif "CCAA" in e and "EMPRESA" not in e:
        datos["CCAA"] = valor
    elif "PROVINCIA" in e and "EMPRESA" not in e:
        datos["PROVINCIA"] = valor
    elif "TELEFONO" in en and "EMPRESA" not in e:
        datos["TELEFONO"] = valor
    elif "EMAIL" in e:
        datos["EMAIL"] = valor
    elif "LOCALIDAD" in e and ("5000" in e or "5.000" in e or "INFERIOR" in e):
        datos["RESIDE LOCALIDAD PEQUENA"] = valor
    elif "PERSONA CON DISCAPACIDAD" in e:
        datos["PERSONA CON DISCAPACIDAD"] = valor
    elif "NIVEL DE ESTUDIOS" in e:
        datos["NIVEL DE ESTUDIOS"] = valor
    elif "TITULACION" in en:
        datos["TITULACION"] = valor
    elif "RELACION CON LA EMPRESA" in en:
        datos["RELACION CON LA EMPRESA"] = valor
    elif "LINKEDIN" in e:
        datos["PERFIL LINKEDIN"] = valor
    elif "COMPLEMENTARIA" in e and "DIGITALIZACION" in en and "GESTION" not in en:
        datos["FORMACION DIGITALIZACION"] = valor
    elif "GESTION" in en and "PROYECTOS" in en:
        datos["FORMACION GESTION PROYECTOS"] = valor
    elif "ANOS DE EXPERIENCIA" in en:
        datos["ANOS EXPERIENCIA LABORAL"] = valor
    elif "EXPERIENCIA LABORAL EN PUESTOS" in e or "EXPERIENCIA EN DIGITALIZACION" in en:
        datos["EXPERIENCIA DIGITALIZACION"] = valor
    elif "SITUACION LABORAL" in en:
        datos["SITUACION LABORAL"] = valor
    elif "NOMBRE EMPRESA" in e or ("NOMBRE" in e and "RAZON SOCIAL" in en):
        datos["NOMBRE EMPRESA"] = valor
    elif "NIF EMPRESA" in e or e.strip() == "NIF":
        datos["NIF EMPRESA"] = valor
    elif "DEPARTAMENTO" in e:
        datos["DEPARTAMENTO EMPRESA"] = valor
    elif "PUESTO" in e or "CARGO" in e:
        datos["PUESTO EMPRESA"] = valor
    elif "ACTIVIDAD DE LA EMPRESA" in e:
        datos["ACTIVIDAD EMPRESA"] = valor
    elif "TAMANO EMPRESA" in en:
        datos["TAMANO EMPRESA"] = valor
    elif "DIRECCION" in en and "EMPRESA" in en:
        datos["DIRECCION EMPRESA"] = valor
    elif "CIUDAD" in e and "EMPRESA" in e:
        datos["CIUDAD EMPRESA"] = valor
    elif "CODIGO POSTAL" in e and "EMPRESA" in e:
        datos["CODIGO POSTAL EMPRESA"] = valor
    elif "CCAA" in e and "EMPRESA" in e:
        datos["CCAA EMPRESA"] = valor
    elif "PROVINCIA" in e and "EMPRESA" in e:
        datos["PROVINCIA EMPRESA"] = valor
    elif "TELEFONO" in en and "EMPRESA" in e:
        datos["TELEFONO EMPRESA"] = valor
    elif "PAGINA WEB" in e or ("WEB" in e and "EMPRESA" in e):
        datos["WEB EMPRESA"] = valor
    elif "ANTIGUEDAD" in en and "EMPRESA" in e:
        datos["ANTIGUEDAD EMPRESA"] = valor
    elif "FACTURACION" in en:
        datos["FACTURACION EMPRESA"] = valor
    elif "AMBITO RURAL" in en:
        datos["AMBITO RURAL EMPRESA"] = valor
    elif "MADUREZ DIGITAL" in en:
        datos["MADUREZ DIGITAL EMPRESA"] = valor
    elif "CANALES DE RELACION" in en:
        datos["CANALES RELACION EMPRESA"] = valor
    elif "PERFIL TIC" in en:
        datos["PERFIL TIC EMPRESA"] = valor
    elif "SOSTENIBILIDAD" in e:
        datos["SOSTENIBILIDAD EMPRESA"] = valor
    elif "PLANES DE TRANSFORMACION" in en or ("POLITICAS" in en and "PLANES" in en):
        datos["PLAN DIGITAL EMPRESA"] = valor
    elif "MAXIMA RESPONSABLE" in en or ("EQUIPO DIRECTIVO" in en and "MUJER" in e):
        datos["MUJER DIRECTIVA EMPRESA"] = valor
    elif "PORCENTAJE DE MUJERES" in en:
        datos["PORCENTAJE MUJERES EMPRESA"] = valor
    elif "MOTIVACION" in en:
        datos["MOTIVACION"] = valor


def _extraer_por_tablas(ruta_pdf, tipo):
    datos = {}
    ultima_clave = None

    with pdfplumber.open(ruta_pdf) as pdf:
        for pagina in pdf.pages:
            tablas = pagina.extract_tables()
            if not tablas:
                continue
            for tabla in tablas:
                for fila in tabla:
                    if not fila or len(fila) < 2:
                        ultima_clave = None
                        continue
                    etq = limpiar_texto(fila[0]) if fila[0] else None
                    val = limpiar_texto(fila[1]) if fila[1] else ""
                    if not val:
                        for celda in fila[2:]:
                            v = limpiar_texto(celda) if celda else ""
                            if v:
                                val = v
                                break
                    if not etq:
                        if ultima_clave == "FORMACION DIGITALIZACION" and val:
                            datos["FORMACION GESTION PROYECTOS"] = val
                        ultima_clave = None
                        continue
                    if not val:
                        ultima_clave = None
                        continue
                    _mapear_campo_tabla(etq, val, datos, tipo)
                    ultima_clave = None
    return datos


# ── ESTRATEGIA 2: COORDENADAS X (pdfplumber) ─────────────────
def _extraer_por_coordenadas(ruta_pdf):
    campos_bruto = {}
    with pdfplumber.open(ruta_pdf) as pdf:
        for page in pdf.pages:
            palabras = page.extract_words()
            filas = {}
            for w in palabras:
                clave_fila = round(w["top"] / 4) * 4
                filas.setdefault(clave_fila, []).append(w)
            etq_pendiente = ""
            for top, palabras_fila in sorted(filas.items()):
                etq_p = [w["text"] for w in palabras_fila if w["x0"] < SEP_X]
                val_p = [w["text"] for w in palabras_fila if w["x0"] >= SEP_X]
                etq_raw = " ".join(etq_p).strip()
                val_raw = " ".join(val_p).strip()
                if etq_raw and val_raw:
                    campos_bruto[normalizar(etq_raw)] = val_raw
                    etq_pendiente = ""
                elif etq_raw and not val_raw:
                    etq_pendiente = etq_raw
                elif not etq_raw and val_raw and etq_pendiente:
                    campos_bruto[normalizar(etq_pendiente)] = val_raw
                    etq_pendiente = ""

    for k in list(campos_bruto.keys()):
        if isinstance(campos_bruto[k], str):
            campos_bruto[k] = re.sub(r"\s+", " ", campos_bruto[k]).strip()
    return campos_bruto


def _coord_es_util(campos_bruto):
    con_valor = sum(1 for v in campos_bruto.values() if v and len(str(v)) > 1)
    clave_ap  = normalizar("PRIMER APELLIDO")
    apellido  = campos_bruto.get(clave_ap, "")
    corrupto  = apellido and (
        len(apellido) > 25
        or sum(1 for c in apellido if not c.isalpha() and c != " ") > 3
    )
    return con_valor >= 5 and not corrupto


def _construir_mapa_canonico():
    mapa = {}
    for mapeo in MAPEO_POR_TIPO.values():
        for campo in mapeo:
            mapa[normalizar(campo)] = campo

    aliases = [
        ("PRIMER APELLIDO",                  "PRIMER APELLIDO"),
        ("SEGUNDO APELLIDO",                 "SEGUNDO APELLIDO"),
        ("NOMBRE",                           "NOMBRE"),
        ("TIPO DE DOCUMENTO",                "TIPO DE DOCUMENTO"),
        ("No DE DOCUMENTO",                  "NUM DE DOCUMENTO"),
        ("NUM DE DOCUMENTO",                 "NUM DE DOCUMENTO"),
        ("NUMERO DE DOCUMENTO",              "NUM DE DOCUMENTO"),
        ("DIRECCION",                        "DIRECCION"),
        ("NOMBRE EMPRESA RAZON SOCIAL",      "NOMBRE EMPRESA"),
        ("TAMANO EMPRESA",                   "TAMANO EMPRESA"),
        ("ANTIGUEDAD DE LA EMPRESA",         "ANTIGUEDAD EMPRESA"),
        ("FACTURACION ULTIMO ANO",           "FACTURACION EMPRESA"),
        ("AMBITO RURAL",                     "AMBITO RURAL EMPRESA"),
        ("NIVEL DE MADUREZ DIGITAL DE LA EMPRESA EN EL", "MADUREZ DIGITAL EMPRESA"),
        ("MOTIVACION PARA CURSAR EL PROGRAMA",           "MOTIVACION"),
        ("DESCRIBIR MOTIVACION PARA CURSAR EL PROGRAMA", "MOTIVACION"),
        ("SITUACION LABORAL ACTUAL",         "SITUACION LABORAL"),
        ("ANOS DE EXPERIENCIA LABORAL",      "ANOS EXPERIENCIA LABORAL"),
        ("PERFIL DE LINKEDIN",               "PERFIL LINKEDIN"),
        ("TELEFONO",                         "TELEFONO"),
        ("DIRECCION EMPRESA",                "DIRECCION EMPRESA"),
        ("CIUDAD EMPRESA",                   "CIUDAD EMPRESA"),
        ("CODIGO POSTAL EMPRESA",            "CODIGO POSTAL EMPRESA"),
        ("CCAA EMPRESA",                     "CCAA EMPRESA"),
        ("PROVINCIA EMPRESA",                "PROVINCIA EMPRESA"),
        ("TELEFONO EMPRESA",                 "TELEFONO EMPRESA"),
        ("PAGINA WEB EMPRESA",               "WEB EMPRESA"),
        ("DEPARTAMENTO EMPRESA",             "DEPARTAMENTO EMPRESA"),
        ("PUESTO EMPRESA",                   "PUESTO EMPRESA"),
        ("PUESTO CARGO",                     "PUESTO EMPRESA"),
        ("RELACION CON LA EMPRESA",          "RELACION CON LA EMPRESA"),
        ("EMPRESA CON POLITICAS DE SOSTENIBILIDAD", "SOSTENIBILIDAD EMPRESA"),
        ("EMPRESA CON POLITICAS O PLANES DE TRANSFORMACION DIGITAL", "PLAN DIGITAL EMPRESA"),
        ("LA MAXIMA RESPONSABLE DE LA EMPRESA", "MUJER DIRECTIVA EMPRESA"),
        ("PORCENTAJE DE MUJERES CON RELACION LABORAL CON LA EMPRESA", "PORCENTAJE MUJERES EMPRESA"),
        ("ACTIVIDAD DE LA EMPRESA CODIGOS CNAE NIVEL LETRA", "ACTIVIDAD EMPRESA"),
        ("FORMACION COMPLEMENTARIA EN DIGITALIZACION",     "FORMACION DIGITALIZACION"),
        ("FORMACION COMPLEMENTARIA EN GESTION DE PROYECTOS", "FORMACION GESTION PROYECTOS"),
        ("EXPERIENCIA LABORAL EN PUESTOS DE DIGITALIZACION", "EXPERIENCIA DIGITALIZACION"),
        ("CANALES DE RELACION DE LA EMPRESA CON CLIENTES Y PROVEEDORES", "CANALES RELACION EMPRESA"),
        ("PROFESIONALES CON PERFIL TIC EN LA EMPRESA",     "PERFIL TIC EMPRESA"),
        ("RESIDE EN UNA LOCALIDAD CON UN NUMERO DE HABITANTES INFERIOR A 5", "RESIDE LOCALIDAD PEQUENA"),
        ("NIF",                              "NIF EMPRESA"),
    ]
    for raw, canonico in aliases:
        mapa[normalizar(raw)] = canonico
    return mapa


_MAPA_CANONICO = _construir_mapa_canonico()


def _normalizado_a_canonico(k_norm):
    if k_norm in _MAPA_CANONICO:
        return _MAPA_CANONICO[k_norm]
    for norm_key, canon in _MAPA_CANONICO.items():
        if k_norm.startswith(norm_key) and len(norm_key) > 5:
            return canon
        if norm_key.startswith(k_norm) and len(k_norm) > 8:
            return canon
    return None


def _mezclar_coordenadas(datos_coord, datos_destino):
    anadidos = 0
    for k_norm, v in datos_coord.items():
        canon = _normalizado_a_canonico(k_norm)
        if canon and canon not in datos_destino:
            datos_destino[canon] = v
            anadidos += 1
    return anadidos


# ── ESTRATEGIA 3: LÍNEAS DE TEXTO ────────────────────────────
_CAMPOS_LINEA = [
    ("PRIMER APELLIDO",                           "PRIMER APELLIDO"),
    ("SEGUNDO APELLIDO",                          "SEGUNDO APELLIDO"),
    ("NOMBRE",                                    "NOMBRE"),
    ("TIPO DE DOCUMENTO",                         "TIPO DE DOCUMENTO"),
    ("Tipo de Documento",                         "TIPO DE DOCUMENTO"),
    ("SEXO",                                      "SEXO"),
    ("FECHA DE NACIMIENTO",                       "FECHA DE NACIMIENTO"),
    ("DIRECCION",                                 "DIRECCION"),
    ("CIUDAD",                                    "CIUDAD"),
    ("CODIGO POSTAL",                             "CODIGO POSTAL"),
    ("CCAA",                                      "CCAA"),
    ("PROVINCIA",                                 "PROVINCIA"),
    ("TELEFONO",                                  "TELEFONO"),
    ("EMAIL",                                     "EMAIL"),
    ("NIVEL DE ESTUDIOS",                         "NIVEL DE ESTUDIOS"),
    ("Nivel de estudios",                         "NIVEL DE ESTUDIOS"),
    ("NOMBRE EMPRESA",                            "NOMBRE EMPRESA"),
    ("NIF EMPRESA",                               "NIF EMPRESA"),
    ("DIRECCION EMPRESA",                         "DIRECCION EMPRESA"),
    ("DIRECCIÓN EMPRESA",                         "DIRECCION EMPRESA"),
    ("DEPARTAMENTO",                              "DEPARTAMENTO EMPRESA"),
    ("TITULACION",                                "TITULACION"),
    ("ANTIGUEDAD DE LA EMPRESA",                  "ANTIGUEDAD EMPRESA"),
    ("AMBITO RURAL",                              "AMBITO RURAL EMPRESA"),
    ("Perfil de LinkedIn",                        "PERFIL LINKEDIN"),
    ("ANOS DE EXPERIENCIA LABORAL",               "ANOS EXPERIENCIA LABORAL"),
    ("SITUACION LABORAL ACTUAL",                  "SITUACION LABORAL"),
    ("Describir motivacion para cursar el programa", "MOTIVACION"),
    ("MOTIVACION para cursar el programa",        "MOTIVACION"),
    ("ACTIVIDAD DE LA EMPRESA",                   "ACTIVIDAD EMPRESA"),
]

_MULTILINEA = {"MOTIVACION", "ACTIVIDAD EMPRESA"}


def _extraer_linea(etiqueta, lineas, multilinea=False):
    patron = re.compile(r"^" + re.escape(etiqueta) + r"\s*(.*)", re.IGNORECASE)
    for i, linea in enumerate(lineas):
        m = patron.match(linea)
        if m:
            valor = m.group(1).strip()
            if multilinea and i + 1 < len(lineas):
                sig = lineas[i + 1]
                sig_norm_first = normalizar(sig).split()[0] if sig.split() else ""
                if sig_norm_first not in ENCABEZADOS_STOP:
                    valor = (valor + " " + sig).strip()
            return valor
    return ""


def _extraer_porcentaje_mujeres(lineas):
    STOP = {"DATOS", "MOTIVACION", "DOCUMENTACION", "ACEPTO", "FIRMA"}
    OPCIONES = re.compile(
        r"inferior a 30\s*%|entre 30\s*%\s*y\s*50\s*%|superior a 50\s*%",
        re.IGNORECASE
    )

    def _norm_op(texto):
        t = texto.strip()
        m = OPCIONES.search(t)
        if m:
            return m.group(0)
        tl = t.lower()
        if "superior" in tl:
            return "superior a 50%"
        if "inferior" in tl:
            return "inferior a 30%"
        if "entre" in tl:
            return "entre 30% y 50%"
        return t

    for i, l in enumerate(lineas):
        if "PORCENTAJE DE MUJERES" in l.upper():
            resto = re.sub(r"PORCENTAJE DE MUJERES[^:]*:?\s*", "", l, flags=re.IGNORECASE).strip()
            resto = OPCIONES.sub("", resto).strip()
            if resto:
                return _norm_op(resto)
            for j in range(i + 1, min(i + 4, len(lineas))):
                sig = lineas[j].strip()
                if any(sig.upper().startswith(k) for k in STOP):
                    break
                sig2 = OPCIONES.sub("", sig).strip()
                if sig2:
                    return _norm_op(sig2)
    return ""


def _extraer_por_lineas(lineas, campos_existentes):
    nuevos = {}
    for etiqueta_pdf, campo_dest in _CAMPOS_LINEA:
        if campo_dest in campos_existentes or campo_dest in nuevos:
            continue
        val = _extraer_linea(etiqueta_pdf, lineas, multilinea=(campo_dest in _MULTILINEA))
        if val:
            nuevos[campo_dest] = re.sub(r"\s+", " ", val).strip()

    if "PORCENTAJE MUJERES EMPRESA" not in campos_existentes:
        porc = _extraer_porcentaje_mujeres(lineas)
        if porc:
            nuevos["PORCENTAJE MUJERES EMPRESA"] = porc

    if "PERFIL LINKEDIN" not in campos_existentes and "PERFIL LINKEDIN" not in nuevos:
        for i, l in enumerate(lineas):
            if re.search(r"perfil\s+de\s+linkedin", l, re.IGNORECASE) and i + 1 < len(lineas):
                val = lineas[i + 1].strip()
                if val and not any(normalizar(val).startswith(k) for k in ENCABEZADOS_STOP):
                    nuevos["PERFIL LINKEDIN"] = val
                break

    if "NUM DE DOCUMENTO" not in campos_existentes and "NUM DE DOCUMENTO" not in nuevos:
        for etq in ("No de Documento", "Nro de Documento", "NUM DE DOCUMENTO",
                    "NUMERO DE DOCUMENTO"):
            val = _extraer_linea(etq, lineas)
            if val:
                nuevos["NUM DE DOCUMENTO"] = val
                break

    if "RESIDE LOCALIDAD PEQUENA" not in campos_existentes:
        for etq in (
            "Reside en una localidad con un numero de habitantes inferior a 5",
            "RESIDE EN UNA LOCALIDAD",
        ):
            val = _extraer_linea(etq, lineas)
            if val:
                nuevos["RESIDE LOCALIDAD PEQUENA"] = val
                break

    if "PERSONA CON DISCAPACIDAD" not in campos_existentes:
        val = _extraer_linea("Persona con discapacidad", lineas)
        if val:
            nuevos["PERSONA CON DISCAPACIDAD"] = val

    return nuevos


# ── ORQUESTADOR: TRIPLE CASCADA ──────────────────────────────
def extraer_campos_pdf(ruta_pdf, tipo, log, lineas=None):
    datos = {}

    # 1. Tablas
    try:
        datos_tablas = _extraer_por_tablas(ruta_pdf, tipo)
        datos.update(datos_tablas)
        log.info(f"  [tablas] {len(datos_tablas)} campos")
    except Exception as e:
        log.warning(f"  [tablas] fallo: {e}")

    # 2. Coordenadas X
    try:
        datos_coord = _extraer_por_coordenadas(ruta_pdf)
        if _coord_es_util(datos_coord):
            anadidos = _mezclar_coordenadas(datos_coord, datos)
            log.info(f"  [coordenadas] {anadidos} campos adicionales")
        else:
            log.warning("  [coordenadas] resultado poco fiable, omitido")
    except Exception as e:
        log.warning(f"  [coordenadas] fallo: {e}")

    # 3. Líneas
    try:
        if lineas is None:
            lineas = leer_lineas_pdf(ruta_pdf)
        datos_lineas = _extraer_por_lineas(lineas, datos)
        datos.update(datos_lineas)
        log.info(f"  [líneas] {len(datos_lineas)} campos adicionales")
    except Exception as e:
        log.warning(f"  [líneas] fallo: {e}")

    log.info(f"  Total campos extraídos: {len(datos)}")
    return datos


# ── LÓGICA EXCEL ─────────────────────────────────────────────
def verificar_hoja_excel(archivo_excel, nombre_hoja_buscar, log):
    try:
        wb = openpyxl.load_workbook(archivo_excel, keep_vba=True, read_only=True)
        hojas = wb.sheetnames
        wb.close()
        for hoja in hojas:
            if hoja.lower() == nombre_hoja_buscar.lower():
                return hoja
        for hoja in hojas:
            if nombre_hoja_buscar.lower() in hoja.lower():
                return hoja
        log.error(f"Hoja '{nombre_hoja_buscar}' no encontrada en la plantilla")
        return None
    except Exception as e:
        log.error(f"Error al leer Excel: {e}")
        return None


def escribir_valor_celda(ws, coordenada, valor):
    try:
        for merged_range in ws.merged_cells.ranges:
            if coordenada in merged_range:
                col_letter, row_num = coordinate_from_string(coordenada)
                col_idx = column_index_from_string(col_letter)
                if col_idx == merged_range.min_col and row_num == merged_range.min_row:
                    break
                else:
                    return False
        ws[coordenada] = valor if not isinstance(valor, str) else str(valor).strip()
        return True
    except Exception:
        return False


def rellenar_excel(datos, tipo, plantilla, salida, nombre_hoja, log):
    mapeo = MAPEO_POR_TIPO[tipo]
    shutil.copy2(plantilla, salida)
    wb = openpyxl.load_workbook(salida, keep_vba=True)
    if nombre_hoja not in wb.sheetnames:
        log.error(f"Hoja '{nombre_hoja}' no encontrada")
        return 0
    ws = wb[nombre_hoja]
    celdas_ok = 0

    for campo, celda in mapeo.items():
        valor = datos.get(campo)
        if valor is None:
            continue
        if escribir_valor_celda(ws, celda, valor):
            celdas_ok += 1

    # Declaraciones siempre "Acepto"
    for campo in DECLARACIONES_COMUNES:
        celda = mapeo.get(campo)
        if celda:
            escribir_valor_celda(ws, celda, "Acepto")
    if tipo != TIPO_AGENTES_1:
        celda = mapeo.get("DECLARO PYME")
        if celda:
            escribir_valor_celda(ws, celda, "Acepto")

    wb.save(salida)
    return celdas_ok


def obtener_nombre_archivo(datos):
    apellido = datos.get("PRIMER APELLIDO", "")
    nombre   = datos.get("NOMBRE", "")
    partes   = [p for p in [apellido, nombre] if p]
    base     = "_".join(partes) if partes else "SinNombre"
    base     = re.sub(r'[\\/*?:"<>|]', "", base)
    return f"Formulario {base}.xlsm"


# ── LOG EN DISCO ─────────────────────────────────────────────
def guardar_log_errores(carpeta, errores):
    if not errores:
        return None
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = carpeta / f"log_errores_{ts}.txt"
    with open(ruta, "w", encoding="utf-8") as f:
        f.write("LOG DE ERRORES - Conversor PDF → Excel\n")
        f.write(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total errores: {len(errores)}\n\n")
        for idx, (pdf_n, tipo_err, detalle) in enumerate(errores, 1):
            f.write(f"{idx}. PDF: {pdf_n}\n")
            f.write(f"   Error: {tipo_err}\n")
            f.write(f"   Detalle: {detalle}\n\n")
    return ruta



# ── COMPROBADOR DE ACTUALIZACIONES ───────────────────────────
def _parsear_version(tag: str) -> tuple:
    """Convierte 'v1.2.3' o '1.2.3' en (1, 2, 3) para comparar."""
    numeros = re.findall(r"\d+", tag)
    return tuple(int(n) for n in numeros[:3]) if numeros else (0,)


def comprobar_actualizacion():
    """
    Consulta la API de GitHub releases y compara con VERSION.
    Devuelve (hay_nueva, version_remota, url_descarga).
    """
    api_url = UPDATE_URL.replace(
        "github.com", "api.github.com/repos"
    )
    resp = requests.get(
        api_url,
        headers={"Accept": "application/vnd.github+json",
                 "User-Agent": "GrupoATU-Updater"},
        timeout=5,
        allow_redirects=True
    )
    resp.raise_for_status()
    data = resp.json()

    tag_remoto   = data.get("tag_name", "")
    url_descarga = data.get("html_url", UPDATE_URL)

    version_remota = _parsear_version(tag_remoto)
    version_actual = _parsear_version(VERSION)

    hay_nueva = version_remota > version_actual
    return hay_nueva, tag_remoto.lstrip("v"), url_descarga


def mostrar_notificacion_actualizacion(version_nueva: str, url_descarga: str):
    """Ventana modal con estilo ATU que informa de la nueva version."""
    ventana = tk.Toplevel()
    ventana.title("Grupo ATU \u2014 Nueva version disponible")
    ventana.geometry("440x260")
    ventana.resizable(False, False)
    ventana.configure(bg=ATU_NAVY)
    ventana.grab_set()
    _set_icon(ventana)

    tk.Frame(ventana, bg=ATU_ORANGE, height=3).pack(fill=tk.X)

    header = tk.Frame(ventana, bg=ATU_NAVY, pady=16)
    header.pack(fill=tk.X)
    tk.Label(header, text="GRUPO ATU", font=("Arial", 16, "bold"),
             bg=ATU_NAVY, fg=ATU_WHITE).pack()
    tk.Label(header, text="Conversor PDF \u2192 Excel",
             font=("Arial", 9), bg=ATU_NAVY, fg=ATU_MID).pack()

    tk.Frame(ventana, bg=ATU_ORANGE, height=3).pack(fill=tk.X)

    body = tk.Frame(ventana, bg=ATU_LIGHT, pady=20, padx=24)
    body.pack(fill=tk.BOTH, expand=True)

    tk.Label(body, text="\U0001f680  Nueva version disponible",
             font=("Arial", 11, "bold"), bg=ATU_LIGHT, fg=ATU_TEXT).pack(pady=(0, 6))
    tk.Label(body,
             text=f"Version actual:      {VERSION}\n"
                  f"Version disponible:  {version_nueva}",
             font=("Consolas", 10), bg=ATU_LIGHT, fg=ATU_TEXT,
             justify=tk.LEFT).pack(pady=(0, 14))

    btn_frame = tk.Frame(body, bg=ATU_LIGHT)
    btn_frame.pack()

    def abrir_y_cerrar():
        webbrowser.open(url_descarga)
        ventana.destroy()

    tk.Button(btn_frame, text="  Descargar actualizacion  ",
              command=abrir_y_cerrar,
              bg=ATU_ORANGE, fg=ATU_WHITE, font=("Arial", 10, "bold"),
              relief=tk.FLAT, cursor="hand2", padx=10, pady=6).pack(side=tk.LEFT, padx=8)
    tk.Button(btn_frame, text="  Continuar sin actualizar  ",
              command=ventana.destroy,
              bg=ATU_BORDER, fg=ATU_TEXT, font=("Arial", 10),
              relief=tk.FLAT, cursor="hand2", padx=10, pady=6).pack(side=tk.LEFT, padx=8)

    tk.Frame(ventana, bg=ATU_ORANGE, height=3).pack(fill=tk.X, side=tk.BOTTOM)
    tk.Label(ventana, text="Hecho por RaulRDA.com \u00b7 Pablo \u00c1lvarez \u00b7 Pelayo Fern\u00e1ndez",
             font=("Arial", 7), bg=ATU_NAVY, fg=ATU_MID).pack(side=tk.BOTTOM, pady=4)

    ventana.wait_window()


def lanzar_comprobacion_actualizacion(root: tk.Tk):
    """
    Ejecuta la comprobacion en un hilo secundario para no bloquear la UI.
    Si hay nueva version, programa la notificacion en el hilo principal.
    Los errores de red se ignoran silenciosamente.
    """
    try:
        hay_nueva, version_nueva, url_descarga = comprobar_actualizacion()
        if hay_nueva:
            root.after(0, lambda: mostrar_notificacion_actualizacion(
                version_nueva, url_descarga
            ))
    except Exception:
        pass  # Sin red o servidor caido -> no interrumpir al usuario


# ── VENTANA SELECTOR ─────────────────────────────────────────
def seleccionar_tipo_plantilla():
    ventana = tk.Toplevel()
    ventana.title("Grupo ATU — Conversor PDF → Excel")
    ventana.geometry("420x280")
    ventana.resizable(False, False)
    ventana.configure(bg=ATU_NAVY)
    ventana.transient()
    ventana.grab_set()
    _set_icon(ventana)

    # Header
    header = tk.Frame(ventana, bg=ATU_NAVY, pady=18)
    header.pack(fill=tk.X)
    tk.Label(header, text="GRUPO ATU", font=("Arial", 18, "bold"),
             bg=ATU_NAVY, fg=ATU_WHITE).pack()
    tk.Label(header, text="Conversor de formularios PDF a Excel",
             font=("Arial", 9), bg=ATU_NAVY, fg=ATU_MID).pack()

    # Separador naranja
    tk.Frame(ventana, bg=ATU_ORANGE, height=3).pack(fill=tk.X)

    # Cuerpo
    body = tk.Frame(ventana, bg=ATU_LIGHT, pady=20)
    body.pack(fill=tk.BOTH, expand=True)

    tk.Label(body, text="Selecciona el tipo de plantilla:",
             font=("Arial", 10, "bold"), bg=ATU_LIGHT, fg=ATU_TEXT).pack(pady=(0, 8))

    tipo_var = tk.StringVar()
    style = ttk.Style()
    style.configure("ATU.TCombobox", fieldbackground=ATU_WHITE, background=ATU_WHITE)
    combobox = ttk.Combobox(body, textvariable=tipo_var,
                            values=list(PLANTILLAS.keys()),
                            state="readonly", width=28, font=("Arial", 10))
    combobox.pack(pady=4)
    combobox.current(0)

    resultado = []

    def aceptar():
        resultado.append(tipo_var.get())
        ventana.destroy()

    def cancelar():
        resultado.append(None)
        ventana.destroy()

    btn_frame = tk.Frame(body, bg=ATU_LIGHT)
    btn_frame.pack(pady=14)

    tk.Button(btn_frame, text="  Aceptar  ", command=aceptar,
              bg=ATU_ORANGE, fg=ATU_WHITE, font=("Arial", 10, "bold"),
              relief=tk.FLAT, cursor="hand2", padx=10, pady=6).pack(side=tk.LEFT, padx=8)
    tk.Button(btn_frame, text="  Cancelar  ", command=cancelar,
              bg=ATU_BORDER, fg=ATU_TEXT, font=("Arial", 10),
              relief=tk.FLAT, cursor="hand2", padx=10, pady=6).pack(side=tk.LEFT, padx=8)

    # Pie
    tk.Frame(ventana, bg=ATU_ORANGE, height=3).pack(fill=tk.X, side=tk.BOTTOM)
    tk.Label(ventana, text="Hecho por RaulRDA.com · Pablo Álvarez · Pelayo Fernández",
             font=("Arial", 7), bg=ATU_NAVY, fg=ATU_MID).pack(side=tk.BOTTOM, pady=4)

    ventana.wait_window()
    return resultado[0] if resultado else None


# ── VENTANA LOG ──────────────────────────────────────────────
class VentanaLog:
    def __init__(self, titulo="Procesando", total=0):
        self.total      = total
        self.procesando = True
        self._actual    = 0

        self.ventana = tk.Toplevel()
        self.ventana.title(f"Grupo ATU — {titulo}")
        self.ventana.geometry("920x640")
        self.ventana.configure(bg=ATU_NAVY)
        self.ventana.transient()
        _set_icon(self.ventana)

        # Header
        header = tk.Frame(self.ventana, bg=ATU_NAVY, pady=10, padx=16)
        header.pack(fill=tk.X)
        tk.Label(header, text="GRUPO ATU", font=("Arial", 14, "bold"),
                 bg=ATU_NAVY, fg=ATU_WHITE).pack(side=tk.LEFT)
        self.lbl_estado = tk.Label(header, text="⏳ Procesando...",
                                   font=("Arial", 9), bg=ATU_NAVY, fg=ATU_ORANGE)
        self.lbl_estado.pack(side=tk.RIGHT)

        tk.Frame(self.ventana, bg=ATU_ORANGE, height=3).pack(fill=tk.X)

        # Barra de progreso
        prog_frame = tk.Frame(self.ventana, bg=ATU_LIGHT, padx=12, pady=6)
        prog_frame.pack(fill=tk.X)
        self.lbl_prog = tk.Label(
            prog_frame, text=f"0 / {total}",
            font=("Arial", 8), bg=ATU_LIGHT, fg=ATU_TEXT
        )
        self.lbl_prog.pack(side=tk.RIGHT)
        self.progress = ttk.Progressbar(
            prog_frame, length=400, mode="determinate",
            maximum=max(total, 1)
        )
        self.progress.pack(fill=tk.X, side=tk.LEFT, expand=True, padx=(0, 8))

        # Área de log
        log_frame = tk.Frame(self.ventana, bg=ATU_LIGHT, padx=12, pady=6)
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.texto = scrolledtext.ScrolledText(
            log_frame, wrap=tk.WORD, font=("Consolas", 9),
            bg="#1E2433", fg="#C8D0E0", insertbackground=ATU_WHITE,
            relief=tk.FLAT, borderwidth=0, padx=8, pady=6
        )
        self.texto.pack(fill=tk.BOTH, expand=True)

        self.texto.tag_config("INFO",    foreground="#8BA7C7")
        self.texto.tag_config("OK",      foreground="#4CAF50")
        self.texto.tag_config("ERROR",   foreground="#EF5350")
        self.texto.tag_config("WARN",    foreground="#FF9800")
        self.texto.tag_config("TITLE",   foreground=ATU_ORANGE, font=("Consolas", 9, "bold"))
        self.texto.tag_config("ARCHIVO", foreground="#E0E0E0", font=("Consolas", 9, "bold"))

        # Barra inferior
        tk.Frame(self.ventana, bg=ATU_ORANGE, height=3).pack(fill=tk.X)
        bottom = tk.Frame(self.ventana, bg=ATU_NAVY, pady=8, padx=12)
        bottom.pack(fill=tk.X)

        tk.Label(bottom, text="Hecho por RaulRDA.com · Pablo Álvarez · Pelayo Fernández",
                 font=("Arial", 7), bg=ATU_NAVY, fg=ATU_MID).pack(side=tk.LEFT)

        self.btn_cerrar = tk.Button(
            bottom, text="  Cerrar  ", command=self.cerrar,
            bg=ATU_BORDER, fg=ATU_TEXT, font=("Arial", 9),
            relief=tk.FLAT, cursor="hand2", pady=4, state=tk.DISABLED
        )
        self.btn_cerrar.pack(side=tk.RIGHT)

        self.ventana.protocol("WM_DELETE_WINDOW", self.cerrar)

    def _escribir(self, texto, tag="INFO"):
        self.texto.insert(tk.END, texto + "\n", tag)
        self.texto.see(tk.END)
        self.ventana.update()
        if sys.stdout:
            print(texto)

    def info(self, msg):    self._escribir(f"  {msg}", "INFO")
    def ok(self, msg):      self._escribir(f"✓ {msg}", "OK")
    def error(self, msg):   self._escribir(f"✗ {msg}", "ERROR")
    def warning(self, msg): self._escribir(f"⚠ {msg}", "WARN")
    def titulo(self, msg):  self._escribir(f"\n{'─'*60}\n  {msg}", "TITLE")
    def archivo(self, msg): self._escribir(f"  📄 {msg}", "ARCHIVO")

    # Alias de compatibilidad
    def log(self, msg, nivel="INFO"): self.info(msg)
    def success(self, msg): self.ok(msg)

    def avanzar(self):
        self._actual += 1
        self.progress["value"] = self._actual
        self.lbl_prog.config(text=f"{self._actual} / {self.total}")
        self.ventana.update()

    def completado(self, fallidos: list):
        self.procesando = False
        self.lbl_estado.config(text="✅ Completado", fg="#4CAF50")
        self._escribir(f"\n{'═'*60}", "TITLE")
        self._escribir("  PROCESO COMPLETADO", "OK")
        if fallidos:
            self._escribir(f"\n  ⚠ No se pudieron convertir estos archivos ({len(fallidos)}):", "WARN")
            for f in fallidos:
                self._escribir(f"    • {f}", "ERROR")
        self._escribir("", "INFO")
        self.btn_cerrar.config(state=tk.NORMAL, bg=ATU_ORANGE, fg=ATU_WHITE,
                               font=("Arial", 9, "bold"))

    def cerrar(self):
        if not self.procesando:
            self.ventana.destroy()


# ── MAIN ─────────────────────────────────────────────────────
def main():
    if sys.stdout and hasattr(sys.stdout, 'buffer'):
        try:
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        except Exception:
            pass

    root = tk.Tk()
    root.withdraw()
    _set_icon(root)

    # Comprobacion de actualizacion en segundo plano
    lanzar_comprobacion_actualizacion(root)

    # 1. Selección del tipo de plantilla (interfaz de main.py)
    tipo = seleccionar_tipo_plantilla()
    if not tipo:
        root.destroy()
        return

    archivo_plantilla, nombre_hoja_buscar, limite = PLANTILLAS[tipo]
    ruta_plantilla = BASE_DIR / archivo_plantilla

    if not ruta_plantilla.exists():
        messagebox.showerror("Error", f"No se encuentra la plantilla:\n{ruta_plantilla}")
        root.destroy()
        return

    # 2. Selección de PDFs
    pdfs = filedialog.askopenfilenames(
        title=f"Selecciona PDFs ({tipo} — máx. {limite})",
        filetypes=[("PDF", "*.pdf")]
    )
    if not pdfs:
        root.destroy()
        return
    if len(pdfs) > limite:
        messagebox.showerror("Error", f"Máximo {limite} PDFs para '{tipo}'")
        root.destroy()
        return

    # 3. Selección de carpeta de salida
    carpeta_str = filedialog.askdirectory(title="Selecciona la carpeta de destino")
    if not carpeta_str:
        root.destroy()
        return
    carpeta_salida = Path(carpeta_str)

    # 4. Verificar hoja Excel una sola vez
    log = VentanaLog(tipo, total=len(pdfs))

    nombre_hoja = verificar_hoja_excel(str(ruta_plantilla), nombre_hoja_buscar, log)
    if not nombre_hoja:
        messagebox.showerror("Error", f"Hoja '{nombre_hoja_buscar}' no encontrada en la plantilla.")
        root.destroy()
        return

    log.info(f"Plantilla : {ruta_plantilla.name}")
    log.info(f"Destino   : {carpeta_salida}")
    log.info(f"PDFs      : {len(pdfs)}")

    exitosos = 0
    fallidos = []   # lista de tuplas (nombre, tipo_error, detalle)

    for i, pdf in enumerate(pdfs, 1):
        pdf_path = Path(pdf)
        log.titulo(f"[{i}/{len(pdfs)}] {pdf_path.name}")

        try:
            # Leer líneas del PDF (para la triple cascada)
            try:
                lineas = leer_lineas_pdf(str(pdf_path))
            except Exception as e:
                raise ValueError(f"PDF ilegible: {e}")

            # Extracción con triple cascada (tablas + coordenadas + líneas)
            datos = extraer_campos_pdf(str(pdf_path), tipo, log, lineas=lineas)

            # Validar que se extrajeron datos reales
            if not any(c in datos for c in CAMPOS_REALES):
                raise ValueError(
                    "No se pudieron extraer campos reales. "
                    "El PDF puede estar escaneado o protegido."
                )

            # Post-procesar campos
            datos = postprocesar_campos(datos, tipo)

            # Generar nombre de archivo de salida
            nombre_salida = obtener_nombre_archivo(datos)
            ruta_salida = carpeta_salida / nombre_salida
            contador = 1
            while ruta_salida.exists():
                ruta_salida = carpeta_salida / f"{ruta_salida.stem}_{contador}.xlsm"
                contador += 1

            # Rellenar Excel
            celdas = rellenar_excel(
                datos, tipo, str(ruta_plantilla), str(ruta_salida), nombre_hoja, log
            )

            if celdas > 0:
                log.ok(f"Generado: {ruta_salida.name}  ({celdas} celdas)")
                exitosos += 1
            else:
                raise ValueError("No se escribió ninguna celda en el Excel")

        except Exception as ex:
            log.error(f"Error: {ex}")
            fallidos.append((pdf_path.name, type(ex).__name__, str(ex)))
            traceback.print_exc()

        log.avanzar()

    # Guardar log de errores en disco si los hay
    ruta_log = None
    if fallidos:
        try:
            ruta_log = guardar_log_errores(carpeta_salida, fallidos)
            if ruta_log:
                log.info(f"Log de errores guardado: {ruta_log.name}")
        except Exception as e:
            log.warning(f"No se pudo guardar el log: {e}")

    log.completado([f for f, *_ in fallidos])

    resumen = f"✅ Convertidos: {exitosos}/{len(pdfs)}"
    if fallidos:
        resumen += f"\n\n⚠ No convertidos ({len(fallidos)}):\n"
        resumen += "\n".join(f"  • {f}" for f, *_ in fallidos)
    resumen += f"\n\nGuardados en:\n{carpeta_salida}"
    if ruta_log:
        resumen += f"\n\nLog de errores:\n{ruta_log}"
    messagebox.showinfo("Resultado", resumen)
    root.destroy()


if __name__ == "__main__":
    main()